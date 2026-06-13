#!/usr/bin/env python3
"""
registro_notas.py — Registro de notas por curso (Liceo Sahagún).

Cruza el CÓDIGO leído en cada hoja OMR con la planilla de estudiantes
(estudiantes/lista_estudiantes.xlsx) y arma un Excel con la nota de cada
alumno, SIN importar el orden en que se escanearon las hojas.

Fuente de las notas: el lector (lector_omr.py) deja, al calificar, un archivo
    profesores/.../<Curso>/resultados/notas_<examen>.json
con un registro por hoja: {archivo, codigo, aciertos, total}.

Salida:
    profesores/.../<Curso>/resultados/registro_<examen>.xlsx

Puede ejecutarse solo (menú) o ser llamado por el lector automáticamente.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ Falta la librería 'openpyxl'. Instálala con:  pip install openpyxl")
    sys.exit(1)

BASE_PROFESORES = Path("profesores")
NOTA_APROBACION = 6.0


# ─────────────────────────────────────────────────────────────────────────────
# Carga de datos
# ─────────────────────────────────────────────────────────────────────────────
def cargar_roster(ruta_curso: Path) -> list[dict]:
    """Lee la planilla del curso → [{n_lista, codigo, nombre}] en orden."""
    archivo = ruta_curso / "estudiantes" / "lista_estudiantes.xlsx"
    if not archivo.exists():
        return []
    wb = load_workbook(archivo)
    ws = wb.active
    roster: list[dict] = []
    for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        nombre = fila[2] if len(fila) > 2 else None
        if not nombre or not str(nombre).strip():
            continue
        codigo = fila[1] if len(fila) > 1 and fila[1] is not None else i
        roster.append({
            "n_lista": fila[0] if fila[0] is not None else i,
            "codigo": str(codigo).strip().zfill(3),
            "nombre": str(nombre).strip(),
        })
    return roster


def cargar_notas(ruta_curso: Path, examen: str) -> list[dict]:
    """Lee las notas que dejó el lector para ese examen."""
    archivo = ruta_curso / "resultados" / f"notas_{examen}.json"
    if not archivo.exists():
        return []
    try:
        return json.loads(archivo.read_text(encoding="utf-8"))
    except Exception:
        return []


def _examenes_disponibles(ruta_curso: Path) -> list[str]:
    carpeta = ruta_curso / "resultados"
    if not carpeta.exists():
        return []
    return sorted(p.stem[len("notas_"):] for p in carpeta.glob("notas_*.json"))


# ─────────────────────────────────────────────────────────────────────────────
# Generación del Excel
# ─────────────────────────────────────────────────────────────────────────────
def generar_registro(ruta_curso: Path, examen: str) -> Optional[Path]:
    """
    Construye registro_<examen>.xlsx cruzando roster + notas.
    Devuelve la ruta del Excel, o None si no hay planilla.
    """
    roster = cargar_roster(ruta_curso)
    if not roster:
        print("  ⚠️  No hay planilla de estudiantes (estudiantes/lista_estudiantes.xlsx).")
        print("     Genérala con gestor_estudiantes.py antes de registrar notas.")
        return None

    notas = cargar_notas(ruta_curso, examen)
    # Indexar notas por código (si un código se escaneó dos veces, gana el último)
    por_codigo: dict[str, dict] = {}
    for r in notas:
        cod = str(r.get("codigo", "")).strip()
        por_codigo[cod] = r

    # ── Construir el libro ────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"

    azul = PatternFill("solid", start_color="1F3864")
    gris = PatternFill("solid", start_color="F2F2F2")
    verde = PatternFill("solid", start_color="E2EFDA")
    rojo = PatternFill("solid", start_color="FCE4E4")
    blanco_negrita = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    negrita = Font(name="Arial", bold=True, size=11)
    normal = Font(name="Arial", size=11)
    centro = Alignment(horizontal="center", vertical="center")
    izq = Alignment(horizontal="left", vertical="center")
    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    titulo = f"REGISTRO DE NOTAS — {examen}   ({ruta_curso.parent.name}°{ruta_curso.name})"
    ws.append([titulo])
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A1"].alignment = izq

    encabezados = ["N° Lista", "Código", "Nombre", "Aciertos", "Total", "Nota", "Estado"]
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        c = ws.cell(row=2, column=col)
        c.fill = azul
        c.font = blanco_negrita
        c.alignment = centro
        c.border = borde

    fila_inicial = 3
    for off, est in enumerate(roster):
        fila = fila_inicial + off
        rec = por_codigo.get(est["codigo"])

        ws.cell(row=fila, column=1, value=est["n_lista"]).alignment = centro
        cod_cell = ws.cell(row=fila, column=2, value=est["codigo"])
        cod_cell.alignment = centro
        cod_cell.number_format = "@"
        cod_cell.font = negrita
        ws.cell(row=fila, column=3, value=est["nombre"]).font = normal
        ws.cell(row=fila, column=3).alignment = izq

        if rec is not None and rec.get("total"):
            aciertos = rec.get("aciertos", 0)
            total = rec.get("total")
            ws.cell(row=fila, column=4, value=aciertos).alignment = centro
            ws.cell(row=fila, column=5, value=total).alignment = centro
            # Nota 0–10 truncada a 1 decimal (no redondea), con piso en 1.0
            cnota = ws.cell(row=fila, column=6,
                            value=f"=MAX(1,TRUNC(10*(D{fila}/E{fila}),1))")
            cnota.alignment = centro
            cnota.font = negrita
            cnota.number_format = "0.0"
            ws.cell(row=fila, column=7,
                    value=f'=IF(F{fila}>={NOTA_APROBACION},"APROBÓ","REPROBÓ")').alignment = centro
        else:
            # No se encontró hoja para ese código
            ws.cell(row=fila, column=6, value="").alignment = centro
            ws.cell(row=fila, column=7, value="No presentó").alignment = centro

        for col in range(1, 8):
            ws.cell(row=fila, column=col).border = borde

    fila_fin = fila_inicial + len(roster) - 1

    # ── Resumen del curso ─────────────────────────────────────────────────────
    fila_prom = fila_fin + 2
    ws.cell(row=fila_prom, column=3, value="Promedio del curso:").font = negrita
    ws.cell(row=fila_prom, column=3).alignment = Alignment(horizontal="right")
    prom = ws.cell(row=fila_prom, column=6,
                   value=f"=IFERROR(TRUNC(AVERAGE(F{fila_inicial}:F{fila_fin}),1),\"-\")")
    prom.font = negrita
    prom.alignment = centro
    prom.number_format = "0.0"

    fila_pres = fila_prom + 1
    ws.cell(row=fila_pres, column=3, value="Presentaron / Total:").font = normal
    ws.cell(row=fila_pres, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=fila_pres, column=6,
            value=f'=COUNT(F{fila_inicial}:F{fila_fin})&" / {len(roster)}"').alignment = centro

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 13
    ws.freeze_panes = "A3"

    # ── Hoja secundaria: hojas cuyo código NO está en la planilla ─────────────
    codigos_roster = {e["codigo"] for e in roster}
    huerfanos = [r for r in notas
                 if str(r.get("codigo", "")).strip() not in codigos_roster]
    if huerfanos:
        ws2 = wb.create_sheet("Por revisar")
        ws2.append(["Archivo escaneado", "Código leído", "Aciertos", "Total"])
        for col in range(1, 5):
            cc = ws2.cell(row=1, column=col)
            cc.fill = azul
            cc.font = blanco_negrita
            cc.alignment = centro
        for r in huerfanos:
            ws2.append([r.get("archivo", ""), str(r.get("codigo", "")),
                        r.get("aciertos", ""), r.get("total", "")])
            ws2.cell(row=ws2.max_row, column=2).number_format = "@"
        ws2.column_dimensions["A"].width = 34
        ws2.column_dimensions["B"].width = 14
        ws2["A" + str(ws2.max_row + 2)] = ("Estos códigos no coinciden con ningún alumno. "
                                           "Revisa que el alumno haya rellenado bien su código.")

    salida = ruta_curso / "resultados_excel" / f"registro_{examen}.xlsx"
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return salida


# ─────────────────────────────────────────────────────────────────────────────
# Menú independiente
# ─────────────────────────────────────────────────────────────────────────────
def _subdirs(ruta: Path) -> list[str]:
    return sorted(d.name for d in ruta.iterdir() if d.is_dir()) if ruta.exists() else []


def _elegir(titulo: str, opciones: list[str], etiqueta: str) -> int:
    print("\n" + "=" * 55)
    print(f"  {titulo}")
    print("=" * 55)
    for i, op in enumerate(opciones, 1):
        print(f"  {i}. {op}")
    while True:
        try:
            v = int(input(f"👉 {etiqueta}: ").strip())
            if 1 <= v <= len(opciones):
                return v - 1
        except ValueError:
            pass
        print("⚠️ Opción inválida.")


def main() -> None:
    print("=" * 55)
    print("   REGISTRO DE NOTAS — LICEO SAHAGÚN")
    print("=" * 55)
    if not BASE_PROFESORES.exists():
        print(f"\n⚠️ No existe '{BASE_PROFESORES}'.")
        sys.exit(1)

    profesores = _subdirs(BASE_PROFESORES)
    ruta_prof = BASE_PROFESORES / profesores[_elegir(
        "Selecciona el PROFESOR", [p.replace('_', ' ') for p in profesores], "Profesor")]
    grados = _subdirs(ruta_prof)
    ruta_grado = ruta_prof / grados[_elegir("Selecciona el GRADO",
                                            [f"Grado {g}" for g in grados], "Grado")]
    cursos = _subdirs(ruta_grado)
    ruta_curso = ruta_grado / cursos[_elegir("Selecciona el CURSO",
                                             [f"Curso {c}" for c in cursos], "Curso")]

    examenes = _examenes_disponibles(ruta_curso)
    if not examenes:
        print("\n⚠️ No hay notas registradas en este curso todavía.")
        print("   Califica primero con lector_omr.py.")
        sys.exit(1)

    examen = examenes[_elegir("Selecciona el EXAMEN", examenes, "Examen")]
    salida = generar_registro(ruta_curso, examen)
    if salida:
        print("\n" + "═" * 55)
        print("✅ Registro de notas generado.")
        print(f"📄 {salida.resolve()}")
        print("═" * 55)


if __name__ == "__main__":
    main()