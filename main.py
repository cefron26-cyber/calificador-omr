#!/usr/bin/env python3
"""
main.py — Calificador OMR (Kivy).

PASO 1: Configuración inicial (una vez) + menú con desplegables.
PASO 2: Estudiantes — lista del curso con código automático.
PASO 3: Exámenes — crear la clave de respuestas y (opcional) el PDF de la hoja.

Reutiliza gestor_salones.py y generador_pdf.py. La clave se guarda en el mismo
formato que lee lector_omr.py / registro_notas.py.
"""

from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.checkbox import CheckBox
from kivy.uix.image import Image
from kivy.core.window import Window
from kivy.core.text import LabelBase
from kivy.graphics import Color, RoundedRectangle, Line
from kivy.factory import Factory
from kivy.metrics import dp
from kivy.clock import Clock
from kivy.lang import Builder
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.widget import Widget
from kivy.utils import platform
from kivy.uix.floatlayout import FloatLayout
from kivy.graphics.texture import Texture

import os
import json
import shutil
import tempfile
from pathlib import Path
import gestor_salones as GS

BASE = GS.BASE_DIR  # carpeta "profesores"

ARCHIVO_LISTA = "lista_estudiantes.xlsx"
ESTUDIANTES_PRUEBA = ["Jesus", "Maria", "Jose"]


# ─────────────────────────────────────────────────────────────────────────────
# TEMA VISUAL · Estilo A "Institucional clásico" (solo apariencia, no lógica)
# ─────────────────────────────────────────────────────────────────────────────
AZUL     = (0.122, 0.220, 0.392, 1)   # #1F3864  (azul del escudo)
AZUL_OSC = (0.078, 0.157, 0.298, 1)   # #14284C  (azul al presionar)
DORADO   = (0.949, 0.718, 0.020, 1)   # #F2B705  (acento dorado)
FONDO    = (0.937, 0.945, 0.961, 1)   # #EFF1F5  (fondo gris claro)
BLANCO   = (1, 1, 1, 1)
# Colores "con significado" para texto con markup (formato hexadecimal):
VERDE = "2E7D32"   # aprobó
ROJO  = "C62828"   # reprobó / error
AMBAR = "F9A825"   # por revisar

_DIR = Path(__file__).resolve().parent


def _registrar_fuentes():
    """Si hay fuentes en assets/fonts, reemplaza la fuente por defecto.
    Detecta el archivo sin importar el nombre exacto (Inter-Regular.ttf,
    Inter_18pt-Regular.ttf, etc.). Si no hay, la app sigue con la letra estándar."""
    base = _DIR / "assets" / "fonts"
    if not base.exists():
        return
    ttfs = list(base.glob("*.ttf"))
    if not ttfs:
        return

    def _buscar(*claves):
        for t in ttfs:
            n = t.name.lower()
            if all(k in n for k in claves):
                return t
        return None

    regular = _buscar("regular") or ttfs[0]
    bold = _buscar("semibold") or _buscar("bold") or regular
    try:
        LabelBase.register(name="Roboto",
                           fn_regular=str(regular),
                           fn_bold=str(bold))
    except Exception:
        pass


def _ruta_logo():
    for r in (_DIR / "assets" / "logo.png", _DIR / "logo.png"):
        if r.exists():
            return str(r)
    return ""


LOGO = _ruta_logo()


def _vibrar(duracion: float = 0.02):
    """Vibración corta SOLO en Android. En PC no hace nada (no rompe la prueba)."""
    try:
        if platform == "android":
            from plyer import vibrator  # type: ignore  # solo existe en Android
            vibrator.vibrate(duracion)
    except Exception:
        pass


def _solicitar_permisos_android():
    """Pide permisos en tiempo de ejecución. Solo corre en Android."""
    try:
        if platform != "android":
            return
        from android.permissions import request_permissions, Permission  # type: ignore  # solo existe en Android
        request_permissions([
            Permission.CAMERA,
            Permission.WRITE_EXTERNAL_STORAGE,
            Permission.READ_EXTERNAL_STORAGE,
        ])
    except Exception:
        pass


class Encabezado(BoxLayout):
    """Barra azul con el escudo, el título y el subtítulo de cada pantalla."""

    def __init__(self, titulo, subtitulo="Liceo Sahagún", **kw):
        super().__init__(orientation="horizontal", size_hint_y=None,
                         height=dp(64), padding=[dp(12), dp(10)],
                         spacing=dp(12), **kw)
        with self.canvas.before:
            Color(rgba=AZUL)
            self._bg = RoundedRectangle(pos=self.pos, size=self.size,
                                        radius=[dp(12)])
        self.bind(pos=self._sync, size=self._sync)

        # Escudo "LS" (cuadro azul oscuro con borde dorado)
        badge = AnchorLayout(size_hint=(None, 1), width=dp(42))
        with badge.canvas.before:
            Color(rgba=(0.086, 0.165, 0.290, 1))
            badge._r = RoundedRectangle(pos=badge.pos, size=badge.size,
                                        radius=[dp(6), dp(6), dp(13), dp(13)])
            Color(rgba=DORADO)
            badge._l = Line(width=1.4, rounded_rectangle=(0, 0, dp(42), dp(42), dp(6)))
        badge.bind(pos=self._badge_sync, size=self._badge_sync)
        self._badge = badge
        badge.add_widget(Label(text="LS", bold=True, color=DORADO, font_size="15sp"))
        self.add_widget(badge)

        # Título + línea dorada + subtítulo
        caja = BoxLayout(orientation="vertical", spacing=dp(2))
        lbl_t = Label(text=titulo, bold=True, color=BLANCO, font_size="18sp",
                      halign="left", valign="bottom", shorten=True)
        sub = Label(text=subtitulo, color=(0.62, 0.71, 0.86, 1), font_size="12sp",
                    halign="left", valign="top")
        for lb in (lbl_t, sub):
            lb.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
        linea = Widget(size_hint=(None, None), size=(dp(46), dp(3)))
        with linea.canvas:
            Color(rgba=DORADO)
            linea._r = RoundedRectangle(pos=linea.pos, size=linea.size, radius=[dp(2)])
        linea.bind(pos=lambda w, *_: setattr(w._r, "pos", w.pos),
                   size=lambda w, *_: setattr(w._r, "size", w.size))
        caja.add_widget(lbl_t)
        caja.add_widget(linea)
        caja.add_widget(sub)
        self.add_widget(caja)

    def _sync(self, *_):
        self._bg.pos = self.pos
        self._bg.size = self.size

    def _badge_sync(self, *_):
        b = self._badge
        b._r.pos = b.pos
        b._r.size = b.size
        b._l.rounded_rectangle = (b.x, b.y, b.width, b.height, dp(6))


class BotonMenu(Button):
    """Botón del menú principal con un ícono vectorial dorado a la izquierda."""

    def __init__(self, icono="", **kw):
        super().__init__(**kw)
        self._icono = icono
        self.halign = "left"
        self.valign = "middle"
        self.bind(pos=self._redibujar, size=self._redibujar)

    def _redibujar(self, *_):
        self.text_size = (self.width - dp(58), self.height)
        self.padding = [dp(50), 0, dp(10), 0]
        self.canvas.after.clear()
        s = dp(20)
        x = self.x + dp(16)
        y = self.center_y - s / 2
        with self.canvas.after:
            Color(rgba=DORADO)
            self._dibujar_icono(self._icono, x, y, s)

    @staticmethod
    def _dibujar_icono(nombre, x, y, s):
        w = 1.6
        if nombre == "estudiantes":
            Line(circle=(x + s * 0.32, y + s * 0.74, s * 0.16), width=w)
            Line(circle=(x + s * 0.70, y + s * 0.74, s * 0.16), width=w)
            Line(rounded_rectangle=(x + s * 0.08, y + s * 0.04,
                                    s * 0.46, s * 0.40, s * 0.12), width=w)
            Line(rounded_rectangle=(x + s * 0.46, y + s * 0.04,
                                    s * 0.46, s * 0.40, s * 0.12), width=w)
        elif nombre == "examenes":
            Line(rounded_rectangle=(x + s * 0.16, y, s * 0.68, s, s * 0.08), width=w)
            Line(points=[x + s * 0.30, y + s * 0.72, x + s * 0.70, y + s * 0.72], width=w)
            Line(points=[x + s * 0.30, y + s * 0.52, x + s * 0.70, y + s * 0.52], width=w)
            Line(points=[x + s * 0.30, y + s * 0.32, x + s * 0.56, y + s * 0.32], width=w)
        elif nombre == "calificar":
            Line(points=[x + s * 0.16, y + s * 0.50, x + s * 0.42, y + s * 0.22,
                         x + s * 0.86, y + s * 0.80], width=2.0)
        elif nombre == "registro":
            Line(rounded_rectangle=(x, y, s, s, s * 0.08), width=w)
            Line(points=[x, y + s * 0.5, x + s, y + s * 0.5], width=w)
            Line(points=[x + s * 0.5, y, x + s * 0.5, y + s], width=w)


class Pildora(Label):
    """Etiqueta con fondo azul claro redondeado (para 'Curso actual')."""

    def __init__(self, **kw):
        super().__init__(**kw)
        self.color = AZUL
        self.font_size = "13sp"
        self.halign = "left"
        self.valign = "middle"
        self.padding = [dp(12), 0, dp(12), 0]
        with self.canvas.before:
            Color(rgba=(0.906, 0.937, 0.984, 1))
            self._r = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(8)])
        self.bind(pos=self._upd, size=self._upd)

    def _upd(self, *_):
        self._r.pos = self.pos
        self._r.size = self.size
        self.text_size = (self.width - dp(24), self.height)


class BotonSalir(Button):
    """Botón minimalista 'Cerrar' con ícono de puerta y flecha (dibujado en canvas)."""

    def __init__(self, **kw):
        super().__init__(**kw)
        self.no_vibra = True  # los botones de cerrar/salir no vibran
        self.halign = "center"
        self.valign = "middle"
        self.bind(pos=self._redibujar, size=self._redibujar)

    def _redibujar(self, *_):
        self.text_size = (self.width, self.height)
        self.padding = [dp(34), 0, dp(8), 0]
        self.canvas.after.clear()
        s = dp(18)
        x = self.x + dp(14)
        y = self.center_y - s / 2
        with self.canvas.after:
            Color(rgba=(0.42, 0.47, 0.55, 1))
            # Marco de la puerta
            Line(rounded_rectangle=(x, y, s * 0.55, s, s * 0.06), width=1.6)
            # Flecha saliendo hacia afuera
            Line(points=[x + s * 0.28, y + s * 0.5, x + s * 1.06, y + s * 0.5], width=1.6)
            Line(points=[x + s * 0.84, y + s * 0.74, x + s * 1.08, y + s * 0.5,
                         x + s * 0.84, y + s * 0.26], width=1.6)


def _caption(texto):
    """Etiqueta pequeña gris que va encima de un campo."""
    lb = Label(text=texto, color=(0.42, 0.47, 0.55, 1), font_size="12sp",
               size_hint_y=None, height=dp(18), halign="left", valign="middle")
    lb.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
    return lb


# Tema visual (embebido para no depender de un archivo .kv externo) -------------
_TEMA_KV = """
#:set C_AZUL      (0.122, 0.220, 0.392, 1)
#:set C_AZUL_OSC  (0.078, 0.157, 0.298, 1)
#:set C_DORADO    (0.949, 0.718, 0.020, 1)
#:set C_DORADO_OSC (0.984, 0.788, 0.290, 1)
#:set C_DORADO_TXT (0.227, 0.169, 0.0, 1)
#:set C_ROJO      (0.780, 0.160, 0.160, 1)
#:set C_ROJO_OSC  (0.620, 0.120, 0.120, 1)
#:set C_VERDE     (0.180, 0.490, 0.196, 1)
#:set C_VERDE_OSC (0.130, 0.370, 0.150, 1)
#:set C_TEXTO     (0.106, 0.129, 0.176, 1)
#:set C_GRIS      (0.42, 0.47, 0.55, 1)
#:set C_BORDE     (0.79, 0.83, 0.89, 1)
#:set C_BLANCO    (1, 1, 1, 1)

<Label>:
    color: C_TEXTO
    font_size: '15sp'

<Button>:
    background_normal: ''
    background_down: ''
    background_disabled_normal: ''
    background_color: 0, 0, 0, 0
    color: C_BLANCO
    font_size: '15sp'
    on_press: app.tap_pulsar(self)
    canvas.before:
        Color:
            rgba: C_AZUL if self.state == 'normal' else C_AZUL_OSC
        RoundedRectangle:
            pos: (self.x + self.width * 0.03, self.y + self.height * 0.03) if self.state == 'down' else self.pos
            size: (self.width * 0.94, self.height * 0.94) if self.state == 'down' else self.size
            radius: [10,]

<BotonAccion@Button>:
    color: C_DORADO_TXT
    bold: True
    canvas.before:
        Color:
            rgba: C_DORADO if self.state == 'normal' else C_DORADO_OSC
        RoundedRectangle:
            pos: (self.x + self.width * 0.03, self.y + self.height * 0.03) if self.state == 'down' else self.pos
            size: (self.width * 0.94, self.height * 0.94) if self.state == 'down' else self.size
            radius: [10,]

<BotonPeligro@Button>:
    color: C_BLANCO
    canvas.before:
        Color:
            rgba: C_ROJO if self.state == 'normal' else C_ROJO_OSC
        RoundedRectangle:
            pos: (self.x + self.width * 0.03, self.y + self.height * 0.03) if self.state == 'down' else self.pos
            size: (self.width * 0.94, self.height * 0.94) if self.state == 'down' else self.size
            radius: [10,]

<BotonExito@Button>:
    color: C_BLANCO
    bold: True
    canvas.before:
        Color:
            rgba: C_VERDE if self.state == 'normal' else C_VERDE_OSC
        RoundedRectangle:
            pos: (self.x + self.width * 0.03, self.y + self.height * 0.03) if self.state == 'down' else self.pos
            size: (self.width * 0.94, self.height * 0.94) if self.state == 'down' else self.size
            radius: [10,]

<BotonOutline@Button>:
    color: C_DORADO_TXT
    bold: True
    canvas.before:
        Color:
            rgba: (1, 1, 1, 1) if self.state == 'normal' else (0.98, 0.96, 0.88, 1)
        RoundedRectangle:
            pos: (self.x + self.width * 0.03, self.y + self.height * 0.03) if self.state == 'down' else self.pos
            size: (self.width * 0.94, self.height * 0.94) if self.state == 'down' else self.size
            radius: [10,]
        Color:
            rgba: C_DORADO
        Line:
            rounded_rectangle: (self.x + self.width * 0.03 + 1, self.y + self.height * 0.03 + 1, self.width * 0.94 - 2, self.height * 0.94 - 2, 10) if self.state == 'down' else (self.x + 1, self.y + 1, self.width - 2, self.height - 2, 10)
            width: 1.4

<BotonSalir>:
    color: C_GRIS
    bold: False
    canvas.before:
        Color:
            rgba: (1, 1, 1, 1) if self.state == 'normal' else (0.95, 0.96, 0.98, 1)
        RoundedRectangle:
            pos: (self.x + self.width * 0.03, self.y + self.height * 0.03) if self.state == 'down' else self.pos
            size: (self.width * 0.94, self.height * 0.94) if self.state == 'down' else self.size
            radius: [10,]
        Color:
            rgba: C_BORDE
        Line:
            rounded_rectangle: (self.x + self.width * 0.03 + 1, self.y + self.height * 0.03 + 1, self.width * 0.94 - 2, self.height * 0.94 - 2, 10) if self.state == 'down' else (self.x + 1, self.y + 1, self.width - 2, self.height - 2, 10)
            width: 1.2

<Spinner>:
    background_normal: ''
    background_down: ''
    background_color: 0, 0, 0, 0
    color: C_AZUL
    font_size: '14sp'
    canvas.before:
        Color:
            rgba: C_BLANCO
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [10,]
    canvas.after:
        Color:
            rgba: C_BORDE
        Line:
            rounded_rectangle: (self.x + 1, self.y + 1, self.width - 2, self.height - 2, 10)
            width: 1.2

<TextInput>:
    background_color: C_BLANCO
    foreground_color: C_TEXTO
    cursor_color: C_DORADO
    hint_text_color: C_GRIS
    font_size: '15sp'
    padding: [12, 12, 12, 12]
    canvas.after:
        Color:
            rgba: C_DORADO if self.focus else C_BORDE
        Line:
            rounded_rectangle: (self.x + 1, self.y + 1, self.width - 2, self.height - 2, 8)
            width: 1.3

<Popup>:
    background_color: 0, 0, 0, 0
    title_color: C_AZUL
    title_size: '17sp'
    separator_color: C_DORADO
    separator_height: '2dp'
    canvas.before:
        Color:
            rgba: C_BLANCO
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [14,]
"""

Builder.load_string(_TEMA_KV)


def _subdirs(p):
    return sorted(d.name for d in p.iterdir() if d.is_dir()) if p.exists() else []


def aviso(titulo, mensaje):
    raiz = FloatLayout()
    cont = BoxLayout(orientation="vertical",
                     padding=[dp(10), dp(40), dp(10), dp(10)], spacing=dp(8))
    lbl = Label(text=str(mensaje), halign="center", valign="top", size_hint_y=None)
    lbl.bind(width=lambda i, w: setattr(i, "text_size", (w, None)))
    lbl.bind(texture_size=lambda i, ts: setattr(i, "height", ts[1]))
    scroll = ScrollView()
    scroll.add_widget(lbl)
    cont.add_widget(scroll)
    raiz.add_widget(cont)
    btn = Factory.BotonPeligro(text="\u2715", size_hint=(None, None),
                               size=(dp(34), dp(34)), pos_hint={"right": 1, "top": 1})
    btn.no_vibra = True
    raiz.add_widget(btn)
    popup = Popup(title=titulo, content=raiz, size_hint=(0.9, 0.6))
    btn.bind(on_release=popup.dismiss)
    popup.open()


def confirmar(titulo, mensaje, on_si, txt_si="S\u00ed, continuar", txt_no="No"):
    """Muestra '¿estás seguro?' con dos opciones. Llama on_si() solo si confirma."""
    cont = BoxLayout(orientation="vertical", padding=dp(12), spacing=dp(12))
    lbl = Label(text=str(mensaje), halign="center", valign="middle")
    lbl.bind(width=lambda i, w: setattr(i, "text_size", (w, None)))
    cont.add_widget(lbl)
    fila = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
    btn_no = BotonSalir(text=txt_no)
    btn_si = Factory.BotonPeligro(text=txt_si)
    fila.add_widget(btn_no)
    fila.add_widget(btn_si)
    cont.add_widget(fila)
    popup = Popup(title=titulo, content=cont, size_hint=(0.85, 0.4),
                  auto_dismiss=False)
    btn_no.bind(on_release=popup.dismiss)

    def _si(*_):
        popup.dismiss()
        try:
            on_si()
        except Exception as e:
            aviso("Error", str(e))
    btn_si.bind(on_release=_si)
    popup.open()


def codigo_de(posicion: int) -> str:
    return f"{posicion:03d}"


def nota_local(aciertos, total):
    """Nota 0-10 truncada a 1 decimal, con piso 1.0 (igual que el lector)."""
    if not total:
        return 0.0
    nota = 10.0 * aciertos / total
    nota = int(nota * 10 + 1e-9) / 10.0
    return max(1.0, nota)


# ─────────────────────────────────────────────────────────────────────────────
# Planilla de estudiantes (mismo formato que registro_notas.py)
# ─────────────────────────────────────────────────────────────────────────────
def leer_lista_estudiantes(ruta_curso):
    archivo = ruta_curso / "estudiantes" / ARCHIVO_LISTA
    if not archivo.exists():
        return []
    from openpyxl import load_workbook
    ws = load_workbook(archivo).active
    nombres = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if len(fila) > 2 and fila[2] and str(fila[2]).strip():
            nombres.append(str(fila[2]).strip())
    return nombres


def guardar_lista_estudiantes(ruta_curso, nombres):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    carpeta = ruta_curso / "estudiantes"
    carpeta.mkdir(parents=True, exist_ok=True)
    archivo = carpeta / ARCHIVO_LISTA
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    ws.append(["N° Lista", "Código", "Nombre"])
    for c in range(1, 4):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center")
    for pos, nombre in enumerate(nombres, start=1):
        ws.append([pos, codigo_de(pos), nombre])
        cod = ws.cell(row=pos + 1, column=2)
        cod.number_format = "@"
        cod.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 34
    wb.save(archivo)
    return archivo


def _openpyxl_disponible():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _nombres_de_filas(filas):
    """Extrae nombres de una hoja (lista de filas). Tolera filas de
    instrucciones arriba: busca el encabezado que contenga 'nombre'."""
    filas = [f for f in filas if f and any(c not in (None, "") for c in f)]
    if not filas:
        return []
    col = None
    inicio = 0
    for i, fila in enumerate(filas[:15]):
        for j, c in enumerate(fila):
            if c is not None and "nombre" in str(c).strip().lower():
                col, inicio = j, i + 1
                break
        if col is not None:
            break
    if col is None:
        ncols = max(len(f) for f in filas)
        mejor, mejor_count = 0, -1
        for i in range(ncols):
            count = 0
            for f in filas:
                if i < len(f) and isinstance(f[i], str):
                    v = f[i].strip()
                    if v and not v.isdigit() and len(v) <= 60 and "\n" not in v:
                        count += 1
            if count > mejor_count:
                mejor, mejor_count = i, count
        col, inicio = mejor, 0
    nombres = []
    for f in filas[inicio:]:
        if col < len(f) and f[col] is not None:
            v = str(f[col]).strip()
            if (v and v.lower() != "nombre"
                    and not v.lower().startswith("nombre")
                    and len(v) <= 60 and "\n" not in v):
                nombres.append(v)
    return nombres


def importar_planilla(ruta_archivo):
    """
    Lee los nombres de los estudiantes desde un .xlsx o .csv.
    Revisa todas las hojas y se queda con la que tenga más nombres.
    Tolera instrucciones arriba (busca el encabezado 'Nombre').
    """
    ruta = Path(ruta_archivo)
    ext = ruta.suffix.lower()
    hojas = []
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(ruta, data_only=True)
        for ws in wb.worksheets:
            hojas.append([list(r) for r in ws.iter_rows(values_only=True)])
    elif ext == ".csv":
        import csv as _csv
        with open(ruta, newline="", encoding="utf-8-sig") as f:
            hojas.append([row for row in _csv.reader(f)])
    else:
        return []

    mejor = []
    for filas in hojas:
        nombres = _nombres_de_filas(filas)
        if len(nombres) > len(mejor):
            mejor = nombres
    return mejor


def generar_plantilla_estudiantes(ruta_curso):
    """Crea un Excel-plantilla para que el profesor llene N° de lista + nombres
    (formato tradicional Apellidos Nombres) y luego lo importe."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    carpeta = ruta_curso / "estudiantes"
    carpeta.mkdir(parents=True, exist_ok=True)
    archivo = carpeta / "plantilla_estudiantes.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    azul = PatternFill("solid", start_color="1F3864")
    blanco = Font(bold=True, color="FFFFFF", size=12)
    centro = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1, value="N\u00b0 Lista")
    ws.cell(row=1, column=2, value="Nombre completo (Apellidos Nombres)")
    for c in (1, 2):
        cell = ws.cell(row=1, column=c)
        cell.fill = azul
        cell.font = blanco
        cell.alignment = centro
    for i in range(1, 41):
        ws.cell(row=i + 1, column=1, value=i).alignment = centro
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 46
    ws.freeze_panes = "A2"

    # Hoja con instrucciones claras
    ins = wb.create_sheet("LÉEME (instrucciones)")
    guia = [
        "CÓMO LLENAR LA PLANILLA DE ESTUDIANTES",
        "",
        "1) Ve a la hoja 'Estudiantes' (pestaña de abajo).",
        "2) Escribe un estudiante por fila en la columna 'Nombre completo'.",
        "3) Usa el formato tradicional: primero APELLIDOS y luego NOMBRES.",
        "      Ejemplo:  Pérez Gómez Juan Carlos",
        "4) El N° de Lista ya viene numerado por orden; puedes agregar más filas.",
        "5) El CÓDIGO de cada alumno (001, 002, 003, ...) lo genera la app sola,",
        "      según el orden de la lista. No tienes que escribirlo.",
        "6) Guarda este archivo (Excel) y vuelve a la app:",
        "      Estudiantes  ->  'Importar planilla'  ->  elige este archivo.",
        "7) La app mostrará cada nombre con su código, para que se los des a los alumnos.",
    ]
    for i, linea in enumerate(guia, start=1):
        celda = ins.cell(row=i, column=1, value=linea)
        if i == 1:
            celda.font = Font(bold=True, size=13)
    ins.column_dimensions["A"].width = 80

    wb.save(archivo)
    return archivo


def abrir_archivo_excel(ruta):
    """Abre/guarda un Excel de forma multiplataforma. En Android lo copia a
    Descargas (visible) y lo abre con una app de hojas de cálculo."""
    if platform == "android":
        _abrir_excel_android(str(ruta))
        return
    try:
        os.startfile(ruta)  # Windows
    except Exception:
        try:
            import subprocess
            subprocess.Popen(["xdg-open", str(ruta)])  # Linux/Mac
        except Exception as e:
            aviso("No se pudo abrir", "Abre el archivo manualmente:\n%s\n\n(%s)"
                  % (ruta, e))


def _abrir_excel_android(ruta):
    """Copia la planilla a Descargas (visible) y la abre con Excel/Sheets."""
    nombre = Path(ruta).name
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    try:
        from jnius import autoclass
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        activity = PythonActivity.mActivity
        VERSION = autoclass("android.os.Build$VERSION")
        with open(ruta, "rb") as f:
            contenido = f.read()
        if VERSION.SDK_INT >= 29:
            ContentValues = autoclass("android.content.ContentValues")
            MediaColumns = autoclass("android.provider.MediaStore$MediaColumns")
            Downloads = autoclass("android.provider.MediaStore$Downloads")
            resolver = activity.getContentResolver()
            values = ContentValues()
            values.put(MediaColumns.DISPLAY_NAME, nombre)
            values.put(MediaColumns.MIME_TYPE, mime)
            values.put(MediaColumns.RELATIVE_PATH, "Download")
            uri = resolver.insert(Downloads.EXTERNAL_CONTENT_URI, values)
            if uri is None:
                raise RuntimeError("No se pudo crear el archivo en Descargas.")
            ostream = resolver.openOutputStream(uri)
            ostream.write(contenido)
            ostream.flush()
            ostream.close()
            try:
                Intent = autoclass("android.content.Intent")
                intent = Intent(Intent.ACTION_VIEW)
                intent.setDataAndType(uri, mime)
                intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                intent.addFlags(Intent.FLAG_ACTIVITY_NEW_TASK)
                activity.startActivity(intent)
            except Exception:
                aviso("Archivo listo",
                      "Se guard\u00f3 en Descargas:\n%s\n\n"
                      "\u00c1brelo con Excel o Google Sheets." % nombre)
        else:
            Environment = autoclass("android.os.Environment")
            carpeta = Environment.getExternalStoragePublicDirectory(
                Environment.DIRECTORY_DOWNLOADS).getAbsolutePath()
            destino = Path(carpeta) / nombre
            with open(destino, "wb") as g:
                g.write(contenido)
            aviso("Archivo listo",
                  "Se guard\u00f3 en Descargas:\n%s\n\n"
                  "\u00c1brelo con Excel o Google Sheets." % nombre)
    except Exception as e:
        aviso("No se pudo abrir", "El archivo est\u00e1 en:\n%s\n\n(%s)" % (ruta, e))


# ─────────────────────────────────────────────────────────────────────────────
# Exámenes (clave JSON + PDF opcional)
# ─────────────────────────────────────────────────────────────────────────────
def listar_examenes(ruta_curso):
    carpeta = ruta_curso / "examenes_claves"
    if not carpeta.exists():
        return []
    out = []
    for f in sorted(carpeta.glob("*.json")):
        try:
            d = json.loads(f.read_text(encoding="utf-8"))
            out.append((f, d.get("examen", f.stem), d.get("total_preguntas", "?")))
        except Exception:
            continue
    return out


def guardar_clave(ruta_curso, nombre, cantidad, clave_letras):
    carpeta = ruta_curso / "examenes_claves"
    carpeta.mkdir(parents=True, exist_ok=True)
    nombre = nombre.strip()
    # El nombre se guarda TAL CUAL (con espacios/acentos); solo el nombre de
    # archivo se hace seguro para el sistema.
    seguro = "".join(c if (c.isalnum() or c in " -_") else "_" for c in nombre)
    seguro = seguro.strip().replace(" ", "_") or "examen"
    datos = {"examen": nombre, "total_preguntas": cantidad,
             "clave_correctas": clave_letras}
    archivo = carpeta / f"respuestas_{seguro}.json"
    archivo.write_text(json.dumps(datos, indent=4, ensure_ascii=False), encoding="utf-8")
    return archivo


def generar_pdf_hoja(ruta_curso, nombre, cantidad):
    import generador_pdf as GP  # importación tardía (necesita reportlab)
    carpeta = ruta_curso / "hojas_pdf"
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta_pdf = carpeta / f"{nombre.replace(' ', '_')}_hoja.pdf"
    GP.generar_hoja_omr(ruta_pdf, cantidad)
    return ruta_pdf


def _guardar_en_descargas(src: Path, filename: str) -> str:
    """
    Copia 'src' a la carpeta de DESCARGAS del dispositivo y devuelve una ruta legible.
    - Android: MediaStore (Android 10+) y, si falla, /storage/emulated/0/Download.
    - iOS: carpeta de documentos de la app (accesible desde la app Archivos).
    - PC: la carpeta Descargas/Downloads del usuario.
    """
    if platform == "android":
        # 1) MediaStore (recomendado en Android moderno)
        try:
            from jnius import autoclass  # type: ignore  # lo provee pyjnius en Android
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            ContentValues = autoclass("android.content.ContentValues")
            MediaColumns = autoclass("android.provider.MediaStore$MediaColumns")
            Downloads = autoclass("android.provider.MediaStore$Downloads")
            activity = PythonActivity.mActivity
            resolver = activity.getContentResolver()
            values = ContentValues()
            values.put(MediaColumns.DISPLAY_NAME, filename)
            values.put(MediaColumns.MIME_TYPE, "application/pdf")
            values.put(MediaColumns.RELATIVE_PATH, "Download")
            uri = resolver.insert(Downloads.EXTERNAL_CONTENT_URI, values)
            stream = resolver.openOutputStream(uri)
            stream.write(src.read_bytes())
            stream.flush()
            stream.close()
            return "Descargas/" + filename
        except Exception:
            pass
        # 2) Copia directa (Android 9 o con almacenamiento heredado)
        destino = Path("/storage/emulated/0/Download") / filename
        destino.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(str(src), str(destino))
        return str(destino)

    if platform == "ios":
        destino = Path(App.get_running_app().user_data_dir) / filename
        shutil.copy(str(src), str(destino))
        return str(destino)

    # PC
    base = None
    if platform == "win" or os.name == "nt":
        # Pregunta a Windows cuál es la carpeta "Descargas" real (respeta OneDrive)
        try:
            import winreg
            sub = r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders"
            guid = "{374DE290-123F-4565-9164-39C4925E467B}"  # known folder: Downloads
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub) as k:
                val, _ = winreg.QueryValueEx(k, guid)
            base = Path(os.path.expandvars(val))
        except Exception:
            base = None
    if base is None or not base.exists():
        base = Path.home() / "Downloads"
        if not base.exists():
            alt = Path.home() / "Descargas"
            base = alt if alt.exists() else base
    base.mkdir(parents=True, exist_ok=True)
    destino = base / filename
    shutil.copy(str(src), str(destino))
    return str(destino)


def generar_pdf_a_descargas(nombre, cantidad) -> str:
    """Genera la hoja PDF en un temporal y la deja SOLO en Descargas. La clave JSON
    se sigue guardando internamente con guardar_clave()."""
    import generador_pdf as GP  # importación tardía (necesita reportlab)
    nombre_arch = f"{nombre.replace(' ', '_')}_hoja.pdf"
    try:
        base_tmp = Path(App.get_running_app().user_data_dir)
    except Exception:
        base_tmp = Path(tempfile.gettempdir())
    tmp = base_tmp / nombre_arch
    GP.generar_hoja_omr(tmp, cantidad)
    destino = _guardar_en_descargas(tmp, nombre_arch)
    try:
        tmp.unlink()
    except Exception:
        pass
    return destino


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: CONFIGURACIÓN INICIAL
# ─────────────────────────────────────────────────────────────────────────────
class SetupScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        root = BoxLayout(orientation="vertical", padding=dp(18), spacing=dp(12))
        root.add_widget(Encabezado("Configuración inicial"))

        intro = Label(text="Esto se hace una sola vez.", size_hint_y=None,
                      height=dp(24), font_size="13sp", color=(0.42, 0.47, 0.55, 1),
                      halign="center", valign="middle")
        intro.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
        root.add_widget(intro)

        root.add_widget(_caption("¿Cómo te llamas?"))
        self.in_nombre = TextInput(hint_text="Nombre y apellido", multiline=False,
                                   size_hint_y=None, height=dp(50))
        root.add_widget(self.in_nombre)

        root.add_widget(_caption("Tus grados y cuántos cursos tiene cada uno"))
        self.filas_box = BoxLayout(orientation="vertical", size_hint_y=None,
                                   spacing=dp(10))
        self.filas_box.bind(minimum_height=self.filas_box.setter("height"))
        scroll = ScrollView()
        scroll.add_widget(self.filas_box)
        root.add_widget(scroll)

        self.filas = []
        self._agregar_fila()

        btn_add = Factory.BotonOutline(text="+ Agregar otro grado", size_hint_y=None,
                                       height=dp(50))
        btn_add.bind(on_release=lambda *_: self._agregar_fila())
        root.add_widget(btn_add)

        btn_crear = Factory.BotonAccion(text="Crear y continuar", size_hint_y=None,
                                        height=dp(56))
        btn_crear.bind(on_release=self._crear)
        root.add_widget(btn_crear)
        self.add_widget(root)

    def _agregar_fila(self):
        fila = BoxLayout(size_hint_y=None, height=dp(74), spacing=dp(12))

        col_g = BoxLayout(orientation="vertical", spacing=dp(4))
        col_g.add_widget(_caption("Grado"))
        in_grado = TextInput(hint_text="ej. 11", multiline=False, input_filter="int",
                             size_hint_y=None, height=dp(48))
        col_g.add_widget(in_grado)

        col_c = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_x=0.55)
        col_c.add_widget(_caption("Cursos"))
        sp = Spinner(text="1", values=[str(i) for i in range(1, 11)],
                     size_hint_y=None, height=dp(48))
        col_c.add_widget(sp)

        fila.add_widget(col_g)
        fila.add_widget(col_c)
        self.filas_box.add_widget(fila)
        self.filas.append((in_grado, sp))

    def _crear(self, *_):
        nombre = self.in_nombre.text.strip()
        if not nombre:
            aviso("Falta el nombre", "Escribe tu nombre para continuar.")
            return
        grados_cursos = []
        for in_grado, sp in self.filas:
            g = in_grado.text.strip()
            if g:
                grados_cursos.append((g, int(sp.text)))
        if not grados_cursos:
            aviso("Faltan grados", "Agrega al menos un grado con sus cursos.")
            return
        GS.inicializar_sistema()
        ruta_prof = GS.crear_profesor(nombre)
        for grado, n in grados_cursos:
            ruta_grado = GS.crear_grado(ruta_prof, grado)
            for i in range(n):
                GS.crear_curso(ruta_grado, chr(65 + i))
        App.get_running_app().ir_a_home()


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: MENÚ PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
class HomeScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.contenedor = BoxLayout(orientation="vertical", padding=dp(14), spacing=dp(8))
        self.add_widget(self.contenedor)

    def on_pre_enter(self, *_):
        self.construir()

    def construir(self):
        self.contenedor.clear_widgets()
        b = self.contenedor
        b.add_widget(Encabezado("Calificador OMR"))

        app = App.get_running_app()
        profesores = _subdirs(BASE)
        if app.profesor and app.profesor in profesores:
            self._prof = app.profesor
        else:
            self._prof = profesores[0] if profesores else ""
        b.add_widget(_caption("Profesor"))
        lbl_prof = Label(text=self._prof.replace("_", " ") or "—",
                         size_hint_y=None, height=dp(40), bold=True,
                         font_size="17sp", halign="left", valign="middle")
        lbl_prof.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
        b.add_widget(lbl_prof)

        fila_gc = BoxLayout(size_hint_y=None, height=dp(70), spacing=dp(10))
        col_g = BoxLayout(orientation="vertical", spacing=dp(4))
        col_g.add_widget(_caption("Grado"))
        self.sp_grado = Spinner(text="", values=[], size_hint_y=None, height=dp(46))
        self.sp_grado.bind(text=lambda *_: self._refrescar_cursos())
        col_g.add_widget(self.sp_grado)
        col_c = BoxLayout(orientation="vertical", spacing=dp(4))
        col_c.add_widget(_caption("Curso"))
        self.sp_curso = Spinner(text="", values=[], size_hint_y=None, height=dp(46))
        self.sp_curso.bind(text=lambda *_: self._guardar_seleccion())
        col_c.add_widget(self.sp_curso)
        fila_gc.add_widget(col_g)
        fila_gc.add_widget(col_c)
        b.add_widget(fila_gc)

        self.lbl_actual = Pildora(text="", size_hint_y=None, height=dp(32))
        b.add_widget(self.lbl_actual)

        b.add_widget(Widget())  # espacio flexible: empuja el botón principal abajo

        btn_cal = Factory.BotonAccion(text="Calificar", size_hint_y=None, height=dp(54))
        btn_cal.bind(on_release=lambda *_: self._entrar())
        b.add_widget(btn_cal)

        btn_add = Factory.BotonOutline(text="+ Agregar grado o curso",
                                       size_hint_y=None, height=dp(46))
        btn_add.bind(on_release=lambda *_: self._popup_agregar())
        b.add_widget(btn_add)

        btn_edit = Factory.BotonOutline(text="Editar / eliminar grados o cursos",
                                        size_hint_y=None, height=dp(46))
        btn_edit.bind(on_release=lambda *_: self._popup_editar())
        b.add_widget(btn_edit)

        self._refrescar_grados()

    def _entrar(self):
        app = App.get_running_app()
        if app.ruta_curso is None:
            aviso("Falta el curso", "Selecciona grado y curso primero.")
            return
        app.sm.current = "dashboard"

    def _refrescar_grados(self):
        prof = self._prof
        grados = _subdirs(BASE / prof) if prof else []
        self.sp_grado.values = grados
        self.sp_grado.text = grados[0] if grados else ""
        self._refrescar_cursos()

    def _refrescar_cursos(self):
        prof, grado = self._prof, self.sp_grado.text
        cursos = _subdirs(BASE / prof / grado) if (prof and grado) else []
        self.sp_curso.values = cursos
        self.sp_curso.text = cursos[0] if cursos else ""
        self._guardar_seleccion()

    def _guardar_seleccion(self):
        app = App.get_running_app()
        app.profesor = self._prof
        app.grado = self.sp_grado.text
        app.curso = self.sp_curso.text
        if app.profesor and app.grado and app.curso:
            app.ruta_curso = BASE / app.profesor / app.grado / app.curso
            self.lbl_actual.text = (f"Curso actual: {app.grado}°{app.curso}"
                                    f"  ·  {app.profesor.replace('_', ' ')}")
        else:
            app.ruta_curso = None
            self.lbl_actual.text = "Selecciona grado y curso."

    def _popup_agregar(self):
        cont = BoxLayout(orientation="vertical", spacing=dp(8), padding=dp(10))
        cont.add_widget(Label(text="Agregar un grado nuevo (con sus cursos)",
                              size_hint_y=None, height=dp(28)))
        fila = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
        fila.add_widget(Label(text="Grado:", size_hint_x=0.35))
        in_g = TextInput(hint_text="ej. 10", multiline=False, input_filter="int")
        fila.add_widget(in_g)
        cont.add_widget(fila)
        fila2 = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
        fila2.add_widget(Label(text="# Cursos:", size_hint_x=0.35))
        sp = Spinner(text="1", values=[str(i) for i in range(1, 11)])
        fila2.add_widget(sp)
        cont.add_widget(fila2)
        btn = Button(text="Crear", size_hint_y=None, height=dp(46))
        cont.add_widget(btn)
        pop = Popup(title="Agregar grado o curso", content=cont, size_hint=(0.9, 0.6))

        def _crear(*_):
            g = in_g.text.strip()
            if not g:
                return
            ruta_prof = BASE / self._prof
            ruta_grado = GS.crear_grado(ruta_prof, g)
            inicio = len(_subdirs(ruta_grado))
            for i in range(int(sp.text)):
                GS.crear_curso(ruta_grado, chr(65 + inicio + i))
            pop.dismiss()
            self.construir()

        btn.bind(on_release=_crear)
        Window.softinput_mode = "pan"
        pop.bind(on_dismiss=lambda *_: setattr(Window, "softinput_mode", ""))
        pop.open()

    def _popup_editar(self):
        prof_dir = BASE / self._prof
        grados = _subdirs(prof_dir)
        cont = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(10))
        cont.add_widget(Label(text="Eliminar grados o cursos", size_hint_y=None,
                              height=dp(28), bold=True))
        cont.add_widget(Label(
            text="Borra lo que sobre. Pide confirmación antes de eliminar.",
            size_hint_y=None, height=dp(22), font_size="12sp",
            color=(0.42, 0.47, 0.55, 1)))
        inner = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(4))
        inner.bind(minimum_height=inner.setter("height"))
        if not grados:
            inner.add_widget(Label(text="No hay grados todavía.",
                                   size_hint_y=None, height=dp(30)))
        for grado in grados:
            fg = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(6))
            lg = Label(text=f"[b]Grado {grado}[/b]", markup=True,
                       halign="left", valign="middle")
            lg.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
            fg.add_widget(lg)
            bg = Factory.BotonPeligro(text="Eliminar grado", size_hint_x=None,
                                      width=dp(140))
            bg.bind(on_release=lambda inst, gr=grado: self._eliminar_grado(gr))
            fg.add_widget(bg)
            inner.add_widget(fg)
            for curso in _subdirs(prof_dir / grado):
                fc = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(6))
                lc = Label(text=f"      Curso {grado}\u00b0{curso}",
                           halign="left", valign="middle")
                lc.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
                fc.add_widget(lc)
                bc = Factory.BotonPeligro(text="X", size_hint_x=None, width=dp(52))
                bc.bind(on_release=lambda inst, gr=grado, cu=curso:
                        self._eliminar_curso(gr, cu))
                fc.add_widget(bc)
                inner.add_widget(fc)
        scroll = ScrollView()
        scroll.add_widget(inner)
        cont.add_widget(scroll)
        btn_cerrar = BotonSalir(text="Cerrar", size_hint_y=None, height=dp(44))
        cont.add_widget(btn_cerrar)
        self._pop_editar = Popup(title="Editar grados/cursos", content=cont,
                                 size_hint=(0.95, 0.85))
        btn_cerrar.bind(on_release=self._pop_editar.dismiss)
        self._pop_editar.open()

    def _eliminar_grado(self, grado):
        def _hacer():
            import shutil
            try:
                shutil.rmtree(BASE / self._prof / grado)
            except Exception as e:
                aviso("Error", str(e))
                return
            try:
                self._pop_editar.dismiss()
            except Exception:
                pass
            self.construir()
            aviso("Listo", f"Se eliminó el grado {grado} y todos sus cursos.")
        confirmar("Eliminar grado",
                  f"¿Eliminar el grado {grado} con TODOS sus cursos, estudiantes y "
                  "notas?\nEsta acción no se puede deshacer.",
                  _hacer, txt_si="Sí, eliminar", txt_no="Cancelar")

    def _eliminar_curso(self, grado, curso):
        def _hacer():
            import shutil
            try:
                shutil.rmtree(BASE / self._prof / grado / curso)
            except Exception as e:
                aviso("Error", str(e))
                return
            try:
                self._pop_editar.dismiss()
            except Exception:
                pass
            self.construir()
            aviso("Listo", f"Se eliminó el curso {grado}\u00b0{curso}.")
        confirmar("Eliminar curso",
                  f"¿Eliminar el curso {grado}\u00b0{curso} con sus estudiantes y "
                  "notas?\nEsta acción no se puede deshacer.",
                  _hacer, txt_si="Sí, eliminar", txt_no="Cancelar")


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: PANEL DEL CURSO (Dashboard)
# ─────────────────────────────────────────────────────────────────────────────
class DashboardScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.contenedor = BoxLayout(orientation="vertical", padding=dp(14), spacing=dp(8))
        self.add_widget(self.contenedor)

    def on_pre_enter(self, *_):
        self.construir()

    def construir(self):
        app = App.get_running_app()
        b = self.contenedor
        b.clear_widgets()

        # Barra superior fija: botón "Cerrar" a la derecha, legible.
        barra_sup = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(6))
        barra_sup.add_widget(Widget(size_hint_x=0.6))
        btn_salir = BotonSalir(text="Cerrar", size_hint_x=0.4)
        btn_salir.bind(on_release=lambda *_: App.get_running_app().stop())
        barra_sup.add_widget(btn_salir)
        b.add_widget(barra_sup)

        b.add_widget(Encabezado("Panel del curso"))

        # Banner persistente "Actualmente estás calificando: Grado - Curso"
        banner = BoxLayout(size_hint_y=None, height=dp(48), padding=[dp(12), dp(6)])
        with banner.canvas.before:
            Color(rgba=DORADO)
            banner._r = RoundedRectangle(pos=banner.pos, size=banner.size,
                                         radius=[dp(10)])
        banner.bind(pos=lambda w, *_: setattr(w._r, "pos", w.pos),
                    size=lambda w, *_: setattr(w._r, "size", w.size))
        lbl = Label(markup=True, color=(0.227, 0.169, 0.0, 1),
                    halign="center", valign="middle", font_size="15sp",
                    text=f"Actualmente estás calificando:  [b]{app.grado}\u00b0 {app.curso}[/b]")
        lbl.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
        banner.add_widget(lbl)
        b.add_widget(banner)

        acciones = {
            "Estudiantes": ("estudiantes", "estudiantes"),
            "Exámenes": ("examenes", "examenes"),
            "Calificar": ("calificar", "calificar"),
            "Registro de notas": ("registro", "registro"),
        }
        for texto, (icono, destino) in acciones.items():
            btn = BotonMenu(icono=icono, text=texto, size_hint_y=None, height=dp(56))
            btn.bind(on_release=lambda inst, d=destino: self._ir(d))
            b.add_widget(btn)

        b.add_widget(Widget())  # espacio flexible

        btn_cambiar = Button(text="\u2190 Cambiar curso", size_hint_y=None, height=dp(46))
        btn_cambiar.bind(on_release=lambda *_: setattr(app.sm, "current", "home"))
        b.add_widget(btn_cambiar)

    def _ir(self, pantalla):
        app = App.get_running_app()
        if app.ruta_curso is None:
            aviso("Falta el curso", "Vuelve y selecciona grado y curso.")
            return
        app.sm.current = pantalla


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: ESTUDIANTES
# ─────────────────────────────────────────────────────────────────────────────
class EstudiantesScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.cont = BoxLayout(orientation="vertical", padding=dp(14), spacing=dp(8))
        self.add_widget(self.cont)

    def on_pre_enter(self, *_):
        # Que el teclado empuje la vista hacia arriba y no tape la caja de texto.
        Window.softinput_mode = "pan"
        self.construir()

    def on_leave(self, *_):
        Window.softinput_mode = ""

    def construir(self):
        app = App.get_running_app()
        self.cont.clear_widgets()
        b = self.cont
        b.add_widget(Encabezado(f"Estudiantes \u00b7 {app.grado}\u00b0{app.curso}"))
        if not _openpyxl_disponible():
            b.add_widget(Label(text="Falta la librería 'openpyxl'.\nInstala: pip install openpyxl"))
            self._boton_volver(b)
            return
        nombres = leer_lista_estudiantes(app.ruta_curso)
        lista_box = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(2))
        lista_box.bind(minimum_height=lista_box.setter("height"))
        if nombres:
            for pos, nombre in enumerate(nombres, start=1):
                lista_box.add_widget(Label(text=f"{codigo_de(pos)}    ·    {nombre}",
                                           size_hint_y=None, height=dp(30)))
        else:
            lista_box.add_widget(Label(text="(Aún no hay estudiantes)",
                                       size_hint_y=None, height=dp(30)))
        scroll = ScrollView()
        scroll.add_widget(lista_box)
        b.add_widget(scroll)
        fila = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(6))
        self.in_nombre = TextInput(hint_text="Nombre del estudiante", multiline=False)
        fila.add_widget(self.in_nombre)
        btn_add = Button(text="Agregar", size_hint_x=0.3)
        btn_add.bind(on_release=self._agregar)
        fila.add_widget(btn_add)
        b.add_widget(fila)
        btn_plantilla = Factory.BotonOutline(text="Descargar plantilla para llenar",
                                             size_hint_y=None, height=dp(44))
        btn_plantilla.bind(on_release=self._descargar_plantilla)
        b.add_widget(btn_plantilla)
        btn_importar = Button(text="Importar planilla (Excel/CSV)",
                              size_hint_y=None, height=dp(44))
        btn_importar.bind(on_release=self._importar)
        b.add_widget(btn_importar)
        btn_vaciar = Factory.BotonPeligro(text="Vaciar lista", size_hint_y=None,
                                          height=dp(40))
        btn_vaciar.bind(on_release=self._vaciar)
        b.add_widget(btn_vaciar)
        self._boton_volver(b)

    def _boton_volver(self, b):
        btn = Button(text="← Volver al menú", size_hint_y=None, height=dp(44))
        btn.bind(on_release=lambda *_: setattr(App.get_running_app().sm, "current", "dashboard"))
        b.add_widget(btn)

    def _agregar(self, *_):
        app = App.get_running_app()
        nombre = self.in_nombre.text.strip()
        if not nombre:
            return
        nombres = leer_lista_estudiantes(app.ruta_curso)
        nombres.append(nombre)
        guardar_lista_estudiantes(app.ruta_curso, nombres)
        self.construir()

    def _descargar_plantilla(self, *_):
        app = App.get_running_app()
        try:
            ruta = generar_plantilla_estudiantes(app.ruta_curso)
        except Exception as e:
            aviso("No se pudo", "No se pudo crear la plantilla:\n%s" % e)
            return
        abrir_archivo_excel(ruta)

    def _vaciar(self, *_):
        app = App.get_running_app()

        def _hacer():
            guardar_lista_estudiantes(app.ruta_curso, [])
            self.construir()
        confirmar("Vaciar lista",
                  "\u00bfSeguro que quieres borrar TODA la lista de estudiantes "
                  "de este curso?\nEsta acci\u00f3n no se puede deshacer.",
                  _hacer)

    _COD_IMPORTAR = 0xE57D

    def _importar(self, *_):
        if platform == "android":
            self._importar_android()
            return
        app = App.get_running_app()
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            ruta = filedialog.askopenfilename(
                title="Elige la planilla (Excel o CSV)",
                filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"), ("Todos", "*.*")])
            root.destroy()
        except Exception as e:
            aviso("Error", f"No se pudo abrir el explorador: {e}")
            return
        if not ruta:
            return
        nombres = importar_planilla(ruta)
        if not nombres:
            aviso("Sin nombres", "No encontré nombres en ese archivo.\n"
                  "Debe tener una columna con los nombres.")
            return
        guardar_lista_estudiantes(app.ruta_curso, nombres)
        self.construir()
        aviso("Listo", f"Se importaron {len(nombres)} estudiantes.")

    def _importar_android(self):
        try:
            from jnius import autoclass  # type: ignore
            from android import activity  # type: ignore
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            ctx = PythonActivity.mActivity
            Intent = autoclass("android.content.Intent")
            intent = Intent(Intent.ACTION_GET_CONTENT)
            intent.setType("*/*")
            intent.addCategory(Intent.CATEGORY_OPENABLE)
            activity.bind(on_activity_result=self._on_importar_result)
            ctx.startActivityForResult(intent, self._COD_IMPORTAR)
        except Exception as e:
            aviso("Importar", "No se pudo abrir el explorador: %s" % e)

    def _on_importar_result(self, request, result, intent):
        if request != self._COD_IMPORTAR:
            return
        try:
            from android import activity  # type: ignore
            activity.unbind(on_activity_result=self._on_importar_result)
        except Exception:
            pass
        if result not in (-1,) or intent is None:
            return
        try:
            uri = intent.getData()
        except Exception:
            uri = None
        if uri is None:
            return
        Clock.schedule_once(lambda dt: self._procesar_importar(uri), 0.1)

    def _procesar_importar(self, uri):
        app = App.get_running_app()
        try:
            ext = self._ext_de_uri(uri)
            destino = Path(app.user_data_dir) / ("planilla_import" + ext)
            ok = self._copiar_uri(uri, destino)
            if not ok:
                aviso("Importar", "No se pudo leer el archivo seleccionado.")
                return
            nombres = importar_planilla(destino)
            try:
                destino.unlink()
            except Exception:
                pass
        except Exception as e:
            import traceback
            aviso("Importar", "Error al leer la planilla:\n%s\n\n%s"
                  % (e, traceback.format_exc()))
            return
        if not nombres:
            aviso("Sin nombres", "No encontr\u00e9 nombres en ese archivo.\n"
                  "Usa la plantilla y llena la columna 'Nombre completo'.")
            return
        guardar_lista_estudiantes(app.ruta_curso, nombres)
        self.construir()
        aviso("Listo", "Se importaron %d estudiantes." % len(nombres))

    def _ext_de_uri(self, uri):
        try:
            from jnius import autoclass  # type: ignore
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            resolver = PythonActivity.mActivity.getContentResolver()
            OpenableColumns = autoclass("android.provider.OpenableColumns")
            cursor = resolver.query(uri, None, None, None, None)
            nombre = ""
            if cursor is not None:
                idx = cursor.getColumnIndex(OpenableColumns.DISPLAY_NAME)
                if cursor.moveToFirst() and idx >= 0:
                    nombre = cursor.getString(idx) or ""
                cursor.close()
            ext = Path(nombre).suffix.lower()
            return ext if ext in (".xlsx", ".xlsm", ".csv") else ".xlsx"
        except Exception:
            return ".xlsx"

    def _copiar_uri(self, uri, destino):
        from jnius import autoclass  # type: ignore
        File = autoclass("java.io.File")
        FileOutputStream = autoclass("java.io.FileOutputStream")
        FileUtils = autoclass("android.os.FileUtils")
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        resolver = PythonActivity.mActivity.getContentResolver()
        try:
            istream = resolver.openInputStream(uri)
            if istream is None:
                return False
            ostream = FileOutputStream(File(str(destino)))
            FileUtils.copy(istream, ostream)
            istream.close()
            ostream.close()
        except Exception:
            return False
        return Path(destino).exists() and Path(destino).stat().st_size > 0


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: EXÁMENES (Paso 3)
# ─────────────────────────────────────────────────────────────────────────────
class EsamenesScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.cont = BoxLayout(orientation="vertical", padding=dp(14), spacing=dp(8))
        self.add_widget(self.cont)

    def on_pre_enter(self, *_):
        self.construir()

    def construir(self):
        app = App.get_running_app()
        self.cont.clear_widgets()
        b = self.cont
        b.add_widget(Encabezado(f"Ex\u00e1menes \u00b7 {app.grado}\u00b0{app.curso}"))
        examenes = listar_examenes(app.ruta_curso)
        lista_box = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(2))
        lista_box.bind(minimum_height=lista_box.setter("height"))
        if examenes:
            for ruta, nombre, npreg in examenes:
                fila = BoxLayout(size_hint_y=None, height=dp(60), spacing=dp(8))
                lbl = Label(
                    text=f"{nombre}\n[size=12sp][color=6E7681]{npreg} preguntas[/color][/size]",
                    markup=True, halign="left", valign="middle")
                lbl.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
                fila.add_widget(lbl)
                btn_dl = Factory.BotonOutline(text="Descargar", size_hint_x=None,
                                              width=dp(120))
                btn_dl.bind(on_release=lambda inst, n=nombre, q=npreg: self._descargar(n, q))
                fila.add_widget(btn_dl)
                btn_x = Factory.BotonPeligro(text="X", size_hint_x=None, width=dp(46))
                btn_x.bind(on_release=lambda inst, r=ruta: self._borrar(r))
                fila.add_widget(btn_x)
                lista_box.add_widget(fila)
        else:
            lista_box.add_widget(Label(text="(Aún no hay exámenes)",
                                       size_hint_y=None, height=dp(30)))
        scroll = ScrollView()
        scroll.add_widget(lista_box)
        b.add_widget(scroll)
        btn_nuevo = Button(text="+ Crear examen nuevo", size_hint_y=None, height=dp(50))
        btn_nuevo.bind(on_release=lambda *_: self._popup_nuevo())
        b.add_widget(btn_nuevo)
        btn_volver = Button(text="← Volver al menú", size_hint_y=None, height=dp(44))
        btn_volver.bind(on_release=lambda *_: setattr(App.get_running_app().sm, "current", "dashboard"))
        b.add_widget(btn_volver)

    def _borrar(self, ruta):
        try:
            ruta.unlink()
        except Exception:
            pass
        self.construir()

    def _descargar(self, nombre, npreg):
        try:
            cantidad = int(npreg)
        except (ValueError, TypeError):
            aviso("No se pudo", "Ese examen no tiene un número de preguntas válido.")
            return
        try:
            dest = generar_pdf_a_descargas(nombre, cantidad)
            aviso("Descargado", f"PDF guardado en:\n{dest}")
        except Exception as e:
            aviso("No se pudo descargar", str(e))

    def _popup_nuevo(self):
        cont = BoxLayout(orientation="vertical", spacing=dp(8), padding=dp(12))
        cont.add_widget(Label(text="Nuevo examen", size_hint_y=None, height=dp(28),
                              font_size="16sp"))
        cont.add_widget(_caption("Nombre del examen:"))
        in_nombre = TextInput(hint_text="ej. Ingles Primer Parcial", multiline=False,
                              size_hint_y=None, height=dp(46))
        cont.add_widget(in_nombre)
        cont.add_widget(_caption("Número de preguntas (1 a 50):"))
        in_num = TextInput(hint_text="ej. 20", multiline=False, input_filter="int",
                           size_hint_y=None, height=dp(46))
        cont.add_widget(in_num)
        lbl_msg = Label(text="", size_hint_y=None, height=dp(24),
                        color=(0.78, 0.16, 0.16, 1))
        cont.add_widget(lbl_msg)
        btn_generar = Factory.BotonAccion(text="Generar claves", size_hint_y=None,
                                          height=dp(50))
        cont.add_widget(btn_generar)
        pop = Popup(title="Crear examen", content=cont, size_hint=(0.92, 0.6))

        def _generar(*_):
            nombre = in_nombre.text.strip()
            if not nombre:
                lbl_msg.text = "Escribe el nombre del examen."
                return
            try:
                cantidad = int(in_num.text)
            except ValueError:
                lbl_msg.text = "El número de preguntas debe ser un número."
                return
            if not (1 <= cantidad <= 50):
                lbl_msg.text = "El número de preguntas debe ir de 1 a 50."
                return
            app = App.get_running_app()

            def _ir():
                pop.dismiss()
                app.sm.get_screen("claves").preparar(nombre, cantidad)
                app.sm.current = "claves"

            existentes = [nom for _, nom, _ in listar_examenes(app.ruta_curso)]
            if nombre in existentes:
                confirmar(
                    "Examen existente",
                    f"Ya existe un examen llamado '{nombre}'.\n"
                    "Si continúas, reemplazarás sus respuestas.",
                    _ir, txt_si="Sí, reemplazar", txt_no="Cancelar")
            else:
                _ir()

        btn_generar.bind(on_release=_generar)
        # Solo aquí: que el teclado suba la pantalla para ver lo que se escribe.
        Window.softinput_mode = "pan"
        pop.bind(on_dismiss=lambda *_: setattr(Window, "softinput_mode", ""))
        pop.open()


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: GENERAR CLAVES (estilo ZipGrade)  ·  un desplegable A/B/C/D por pregunta
# ─────────────────────────────────────────────────────────────────────────────
class ClavesScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.cont = BoxLayout(orientation="vertical", padding=dp(12), spacing=dp(8))
        self.add_widget(self.cont)
        self.nombre = ""
        self.cantidad = 0
        self.spinners = []

    def preparar(self, nombre, cantidad):
        self.nombre = nombre
        self.cantidad = int(cantidad)

    def on_pre_enter(self, *_):
        self.construir()

    def construir(self):
        self.cont.clear_widgets()
        b = self.cont
        b.add_widget(Encabezado("Generando claves"))
        b.add_widget(Label(
            text="[b]Examen:[/b] %s    \u00b7    %d preguntas" % (self.nombre, self.cantidad),
            markup=True, size_hint_y=None, height=dp(28)))
        b.add_widget(Label(
            text="Elige la respuesta correcta de cada pregunta.",
            size_hint_y=None, height=dp(22), font_size="12sp",
            color=(0.42, 0.47, 0.55, 1)))

        n = self.cantidad
        self.spinners = [None] * (n + 1)  # 1-indexado
        if n <= 10:
            columnas = [list(range(1, n + 1))]
        else:
            izq = (n + 1) // 2  # impar: el lado izquierdo lleva una más
            columnas = [list(range(1, izq + 1)), list(range(izq + 1, n + 1))]

        fila_cols = BoxLayout(orientation="horizontal", spacing=dp(14),
                              size_hint_y=None)
        fila_cols.bind(minimum_height=fila_cols.setter("height"))
        for col in columnas:
            colbox = BoxLayout(orientation="vertical", spacing=dp(6),
                               size_hint_y=None)
            colbox.bind(minimum_height=colbox.setter("height"))
            for q in col:
                fila = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(8))
                etq = Label(text="%d." % q, size_hint_x=None, width=dp(42),
                            halign="right", valign="middle", bold=True)
                etq.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
                fila.add_widget(etq)
                sp = Spinner(text="A", values=("A", "B", "C", "D"))
                self.spinners[q] = sp
                fila.add_widget(sp)
                colbox.add_widget(fila)
            fila_cols.add_widget(colbox)

        scroll = ScrollView()
        scroll.add_widget(fila_cols)
        b.add_widget(scroll)

        fila_chk = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        self.chk_pdf = CheckBox(size_hint_x=None, width=dp(36), active=True)
        fila_chk.add_widget(self.chk_pdf)
        cap = Label(text="Guardar tambi\u00e9n el PDF de la hoja en Descargas",
                    halign="left", valign="middle")
        cap.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
        fila_chk.add_widget(cap)
        b.add_widget(fila_chk)

        fila_btn = BoxLayout(size_hint_y=None, height=dp(54), spacing=dp(10))
        btn_cancel = BotonSalir(text="Cancelar")
        btn_cancel.bind(on_release=lambda *_: setattr(
            App.get_running_app().sm, "current", "examenes"))
        btn_guardar = Factory.BotonAccion(text="Guardar claves")
        btn_guardar.bind(on_release=self._guardar)
        fila_btn.add_widget(btn_cancel)
        fila_btn.add_widget(btn_guardar)
        b.add_widget(fila_btn)

    def _guardar(self, *_):
        app = App.get_running_app()
        clave = [self.spinners[q].text for q in range(1, self.cantidad + 1)]
        guardar_clave(app.ruta_curso, self.nombre, self.cantidad, clave)
        mensaje = "Examen '%s' guardado con %d respuestas." % (self.nombre, self.cantidad)
        if getattr(self, "chk_pdf", None) and self.chk_pdf.active:
            try:
                generar_pdf_a_descargas(self.nombre, self.cantidad)
                mensaje += "\n\nPDF de la hoja guardado en Descargas."
            except Exception as e:
                mensaje += "\n\n(No se pudo guardar el PDF: %s)" % e
        app.sm.current = "examenes"
        aviso("Listo", mensaje)


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: CALIFICAR (Paso 4)  ·  digitalizar -> leer -> guardar notas
# ─────────────────────────────────────────────────────────────────────────────
class CalificarScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.cont = BoxLayout(orientation="vertical", padding=dp(14), spacing=dp(8))
        self.add_widget(self.cont)
        self.resultados = []
        self.examen = ""
        self._examen_sel = ""

    def on_pre_enter(self, *_):
        self.resultados = []
        self.examen = ""
        self._examen_sel = ""
        self.construir()

    def construir(self):
        app = App.get_running_app()
        self.cont.clear_widgets()
        b = self.cont
        b.add_widget(Encabezado(f"Calificar \u00b7 {app.grado}\u00b0{app.curso}"))

        examenes = listar_examenes(app.ruta_curso)
        if not examenes:
            b.add_widget(Label(text="Primero crea un examen en 'Exámenes'."))
            self._volver(b)
            return

        self._mapa_claves = {nombre: ruta for ruta, nombre, _ in examenes}
        fila = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
        fila.add_widget(Label(text="Examen:", size_hint_x=0.3))
        nombres = list(self._mapa_claves.keys())
        inicial = self._examen_sel if self._examen_sel in nombres else nombres[0]
        self._examen_sel = inicial
        self.sp_examen = Spinner(text=inicial, values=nombres)
        self.sp_examen.bind(text=self._cambio_examen)
        fila.add_widget(self.sp_examen)
        b.add_widget(fila)

        btn = Factory.BotonAccion(text="Cargar foto(s) y calificar", size_hint_y=None,
                                  height=dp(52))
        btn.bind(on_release=self._calificar)
        b.add_widget(btn)

        btn_cam = Button(text="Tomar foto con la cámara", size_hint_y=None, height=dp(48))
        btn_cam.bind(on_release=self._tomar_foto)
        b.add_widget(btn_cam)

        lista = BoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(2))
        lista.bind(minimum_height=lista.setter("height"))
        if self.resultados:
            lista.add_widget(Label(text="[b]Resultados[/b]", markup=True,
                                   size_hint_y=None, height=dp(26)))
            for r in self.resultados:
                codigo = r.get("codigo", "???")
                nombre = r.get("nombre", "")
                dudoso = ("?" in codigo) or codigo in ("", "???", "ERROR")
                if codigo == "ERROR":
                    fila_r = BoxLayout(size_hint_y=None, height=dp(34))
                    fila_r.add_widget(Label(
                        text=f"[color={ROJO}]{r['archivo']}: error al leer[/color]",
                        markup=True))
                    lista.add_widget(fila_r)
                    continue
                n = nota_local(r["aciertos"], r["total"])
                col = VERDE if n >= 6.0 else ROJO
                if dudoso or not nombre:
                    # Código no identificado: avisar y ofrecer corregir
                    fila_r = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
                    etq = nombre or f"Código {codigo}"
                    lb = Label(text=f"[color={AMBAR}]{etq}  ·  revisar código[/color]",
                               markup=True, halign="left", valign="middle")
                    lb.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
                    fila_r.add_widget(lb)
                    btn_e = Button(text="Corregir", size_hint_x=None, width=dp(110))
                    btn_e.bind(on_release=lambda inst, rr=r: self._editar_codigo(rr))
                    fila_r.add_widget(btn_e)
                    lista.add_widget(fila_r)
                else:
                    lista_num = str(int(codigo)) if codigo.isdigit() else codigo
                    fila_r = BoxLayout(size_hint_y=None, height=dp(34))
                    texto = (f"[b]{lista_num}. {nombre}[/b]"
                             f"     \u00b7     nota [color={col}][b]{n:.1f}[/b][/color]")
                    lb = Label(text=texto, markup=True, halign="left", valign="middle")
                    lb.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
                    fila_r.add_widget(lb)
                    lista.add_widget(fila_r)
        else:
            lista.add_widget(Label(text="Elige el examen y carga las fotos.",
                                   size_hint_y=None, height=dp(28)))
        scroll = ScrollView()
        scroll.add_widget(lista)
        b.add_widget(scroll)

        self._volver(b)

    def _cambio_examen(self, inst, val):
        if val and val != self._examen_sel:
            self._examen_sel = val
            self.resultados = []   # los resultados eran de otro examen
            self.construir()

    def _volver(self, b):
        btn = Button(text="← Volver al menú", size_hint_y=None, height=dp(44))
        btn.bind(on_release=lambda *_: setattr(App.get_running_app().sm, "current", "dashboard"))
        b.add_widget(btn)

    def _elegir_archivos(self):
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            rutas = filedialog.askopenfilenames(
                title="Elige la(s) foto(s) de las hojas",
                filetypes=[("Imágenes", "*.png *.jpg *.jpeg *.bmp"), ("Todos", "*.*")])
            root.destroy()
            return list(rutas)
        except Exception as e:
            aviso("Error", f"No se pudo abrir el explorador: {e}")
            return []

    def _calificar(self, *_):
        if platform == "android":
            self._cargar_galeria_android()
            return
        app = App.get_running_app()
        archivos = self._elegir_archivos()
        if not archivos:
            return

        import lector_omr as L
        ruta_curso = app.ruta_curso
        clave_json = self._mapa_claves[self.sp_examen.text]

        # 1) Copiar las fotos a examenes_crudos (igual que la consola)
        dir_crudos = ruta_curso / "examenes_crudos"
        dir_crudos.mkdir(parents=True, exist_ok=True)
        for a in archivos:
            try:
                shutil.copy(a, dir_crudos / Path(a).name)
            except Exception:
                pass

        # 2) DIGITALIZAR: alinear crudos -> procesados (usa digitalizador.py)
        L.auto_digitalizar(ruta_curso)

        # 3) Calificar todas las hojas procesadas
        dir_imgs = ruta_curso / "examenes_procesados"
        dir_salida = ruta_curso / "resultados"
        dir_salida.mkdir(parents=True, exist_ok=True)
        imagenes = L._listar_imagenes(dir_imgs)
        if not imagenes:
            aviso("Sin imágenes", "No se generaron imágenes alineadas.\n"
                  "¿Está 'digitalizador.py' en la carpeta?")
            return

        datos = L.cargar_datos_examen(clave_json)
        examen = datos.nombre
        nuevos = []
        for img in imagenes:
            try:
                res = L._procesar_con_salida(img, clave_json, dir_salida)
                aciertos = sum(1 for p, r in res.respuestas.items()
                               if r == datos.clave.get(p))
                nuevos.append({"archivo": img.name, "codigo": res.codigo_estudiante,
                               "aciertos": aciertos, "total": datos.total_preguntas})
            except Exception as e:
                nuevos.append({"archivo": img.name, "codigo": "ERROR",
                               "aciertos": 0, "total": datos.total_preguntas,
                               "_err": str(e)})

        # 4) Guardar notas_<examen>.json (merge por archivo, igual que la consola)
        notas_path = dir_salida / f"notas_{examen}.json"
        registros = {}
        if notas_path.exists():
            try:
                for r in json.loads(notas_path.read_text(encoding="utf-8")):
                    registros[r.get("archivo", "")] = r
            except Exception:
                pass
        for r in nuevos:
            registros[r["archivo"]] = {k: v for k, v in r.items() if not k.startswith("_")}
        notas_path.write_text(
            json.dumps(list(registros.values()), indent=2, ensure_ascii=False),
            encoding="utf-8")

        # 5) Guardar para mostrar y permitir corrección de código
        self.examen = examen
        self.resultados = nuevos
        self.construir()

    def _editar_codigo(self, r):
        cont = BoxLayout(orientation="vertical", spacing=dp(8), padding=dp(10))
        cont.add_widget(Label(text=f"Hoja: {r['archivo']}", size_hint_y=None, height=dp(24)))
        cont.add_widget(Label(text="Código correcto (3 dígitos):",
                              size_hint_y=None, height=dp(24)))
        actual = "".join(c for c in r.get("codigo", "") if c.isdigit())
        ti = TextInput(text=actual, multiline=False, input_filter="int",
                       size_hint_y=None, height=dp(46))
        cont.add_widget(ti)
        lbl = Label(text="", size_hint_y=None, height=dp(22), color=(0.78, 0.16, 0.16, 1))
        cont.add_widget(lbl)
        btn = Factory.BotonAccion(text="Guardar c\u00f3digo", size_hint_y=None, height=dp(46))
        cont.add_widget(btn)
        pop = Popup(title="Corregir código", content=cont, size_hint=(0.85, 0.55))

        def _ok(*_):
            nuevo = ti.text.strip()
            if not nuevo.isdigit():
                lbl.text = "Escribe solo números."
                return
            nuevo = nuevo.zfill(3)
            r["codigo"] = nuevo
            self._actualizar_notas_codigo(r["archivo"], nuevo)
            pop.dismiss()
            self.construir()

        btn.bind(on_release=_ok)
        pop.open()

    def _actualizar_notas_codigo(self, archivo, nuevo):
        app = App.get_running_app()
        notas_path = app.ruta_curso / "resultados" / f"notas_{self.examen}.json"
        if not notas_path.exists():
            return
        try:
            data = json.loads(notas_path.read_text(encoding="utf-8"))
        except Exception:
            return
        for e in data:
            if e.get("archivo") == archivo:
                e["codigo"] = nuevo
        notas_path.write_text(json.dumps(data, indent=2, ensure_ascii=False),
                              encoding="utf-8")

    # ── Cámara (solo Android / iOS) ───────────────────────────────────────────
    def _ruta_captura(self):
        """Carpeta donde la cámara guardará la foto. En Android usa la carpeta
        externa de la app (sí es accesible por la app de cámara); en otro caso,
        el directorio de datos del usuario."""
        if platform == "android":
            try:
                from jnius import autoclass  # type: ignore
                PythonActivity = autoclass("org.kivy.android.PythonActivity")
                ctx = PythonActivity.mActivity
                ext = ctx.getExternalFilesDir(None)
                if ext is not None:
                    return Path(ext.getAbsolutePath()) / "captura_omr.jpg"
            except Exception:
                pass
        return Path(App.get_running_app().user_data_dir) / "captura_omr.jpg"

    def _tomar_foto(self, *_):
        if platform == "android":
            self._abrir_camara_c4k()
            return
        if platform == "ios":
            try:
                from plyer import camera  # type: ignore
            except Exception:
                aviso("Cámara", "No se pudo acceder a la cámara en este dispositivo.")
                return
            destino = self._ruta_captura()
            try:
                destino.parent.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            try:
                camera.take_picture(filename=str(destino), on_complete=self._foto_lista)
            except Exception as e:
                aviso("Cámara", f"No se pudo abrir la cámara: {e}")
            return
        aviso("Cámara", "Función solo disponible en dispositivos android o ios.")

    def _abrir_camara_c4k(self):
        """Cámara integrada con camera4kivy (CameraX), como las apps pro:
        vista previa en vivo dentro de la app + botón de captura."""
        try:
            from camera4kivy import Preview  # type: ignore
        except Exception as e:
            aviso("Cámara", "No se pudo iniciar la cámara integrada: %s\n\n"
                  "Usa el botón 'Cargar foto(s) y calificar' como alternativa." % e)
            return
        try:
            cont = FloatLayout()
            self._preview = Preview(aspect_ratio="4:3",
                                    size_hint=(1, 1), pos_hint={"x": 0, "y": 0})
            cont.add_widget(self._preview)

            btn_cap = Factory.BotonAccion(text="Capturar", size_hint=(None, None),
                                          size=(dp(180), dp(60)),
                                          pos_hint={"center_x": 0.5, "y": 0.03})
            btn_cap.bind(on_release=lambda *_: self._preview.capture_photo(location="private"))
            cont.add_widget(btn_cap)

            btn_x = Factory.BotonPeligro(text="X", size_hint=(None, None),
                                         size=(dp(54), dp(54)),
                                         pos_hint={"right": 0.98, "top": 0.98})
            cont.add_widget(btn_x)

            self._popup_cam = Popup(title="Tomar foto de la hoja", content=cont,
                                    size_hint=(1, 1), auto_dismiss=False)
            btn_x.bind(on_release=lambda *_: self._popup_cam.dismiss())
            self._popup_cam.bind(
                on_open=lambda *_: self._preview.connect_camera(
                    camera_id="back", filepath_callback=self._foto_c4k,
                    enable_video=False),
                on_pre_dismiss=lambda *_: self._desconectar_c4k())
            self._popup_cam.open()
        except Exception as e:
            import traceback
            aviso("Cámara", "Error iniciando la cámara: %s\n\n%s"
                  % (e, traceback.format_exc()))

    def _desconectar_c4k(self):
        try:
            self._preview.disconnect_camera()
        except Exception:
            pass

    def _foto_c4k(self, path):
        """camera4kivy llama esto cuando la foto quedó guardada (otro hilo)."""
        Clock.schedule_once(lambda dt: self._tras_captura_c4k(path), 0)

    def _tras_captura_c4k(self, path):
        p = Path(str(path)) if path else None
        if p is None or not p.exists():
            return  # puede ser un mensaje de aviso, no una ruta real
        try:
            self._popup_cam.dismiss()
        except Exception:
            pass
        import cv2
        img = cv2.imread(str(p))
        try:
            p.unlink()
        except Exception:
            pass
        if img is None:
            aviso("Cámara", "No se pudo leer la foto capturada.")
            return
        self._calificar_imagen(img)

    # Código de petición para identificar el resultado de la cámara
    _COD_CAMARA = 0xC0FE

    def _tomar_foto_android(self):
        """Abre la cámara con el Intent nativo. Usa MediaStore para crear un
        destino donde la app de cámara SÍ puede escribir en Android 11+."""
        try:
            from jnius import autoclass, cast  # type: ignore
            from android import activity  # type: ignore
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            ctx = PythonActivity.mActivity
            Intent = autoclass("android.content.Intent")
            MediaStore = autoclass("android.provider.MediaStore")
            ContentValues = autoclass("android.content.ContentValues")
            Images = autoclass("android.provider.MediaStore$Images$Media")
            VERSION = autoclass("android.os.Build$VERSION")
            Integer = autoclass("java.lang.Integer")
        except Exception as e:
            aviso("Cámara", "No se pudo acceder a la cámara: %s" % e)
            return
        try:
            resolver = ctx.getContentResolver()
            valores = ContentValues()
            valores.put("_display_name", "captura_omr_tmp.jpg")
            valores.put("mime_type", "image/jpeg")
            if VERSION.SDK_INT >= 29:
                valores.put("relative_path", "Pictures")
            uri = resolver.insert(Images.EXTERNAL_CONTENT_URI, valores)
            if uri is None:
                aviso("Cámara", "No se pudo preparar el almacenamiento de la foto.")
                return
            self._foto_uri = uri
            intent = Intent(MediaStore.ACTION_IMAGE_CAPTURE)
            intent.putExtra(MediaStore.EXTRA_OUTPUT, cast("android.os.Parcelable", uri))
            intent.addFlags(Intent.FLAG_GRANT_WRITE_URI_PERMISSION)
            intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
            activity.bind(on_activity_result=self._on_cam_result)
            ctx.startActivityForResult(intent, self._COD_CAMARA)
        except Exception as e:
            try:
                activity.unbind(on_activity_result=self._on_cam_result)
            except Exception:
                pass
            aviso("Cámara", "No se pudo abrir la cámara: %s" % e)

    def _on_cam_result(self, request, result, intent):
        if request != self._COD_CAMARA:
            return
        try:
            from android import activity  # type: ignore
            activity.unbind(on_activity_result=self._on_cam_result)
        except Exception:
            pass
        Clock.schedule_once(lambda dt: self._procesar_captura_uri(result), 0.4)

    def _eliminar_uri(self, resolver, uri):
        try:
            if uri is not None:
                resolver.delete(uri, None, None)
        except Exception:
            pass

    def _leer_uri_imagen(self, uri):
        """Copia una URI content:// a un archivo temporal y la lee con OpenCV.
        Devuelve (img, bytes_copiados); img es None si no se pudo."""
        from jnius import autoclass  # type: ignore
        File = autoclass("java.io.File")
        FileOutputStream = autoclass("java.io.FileOutputStream")
        FileUtils = autoclass("android.os.FileUtils")
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        resolver = PythonActivity.mActivity.getContentResolver()
        destino = Path(App.get_running_app().user_data_dir) / "captura_omr.jpg"
        copiados = 0
        try:
            istream = resolver.openInputStream(uri)
            if istream is not None:
                ostream = FileOutputStream(File(str(destino)))
                copiados = FileUtils.copy(istream, ostream)
                try:
                    istream.close()
                    ostream.close()
                except Exception:
                    pass
        except Exception:
            return None, 0
        if not destino.exists() or destino.stat().st_size == 0:
            return None, copiados
        import cv2
        img = cv2.imread(str(destino))
        try:
            destino.unlink()
        except Exception:
            pass
        return img, copiados

    def _procesar_captura_uri(self, result):
        """Lee la foto de la cámara (aunque devuelva 'cancelado', pues algunas
        cámaras guardan la imagen igual) y la califica, con diagnóstico."""
        pasos = []
        try:
            from jnius import autoclass  # type: ignore
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            resolver = PythonActivity.mActivity.getContentResolver()
            uri = getattr(self, "_foto_uri", None)
            pasos.append("Código de resultado: %s  (OK = -1)" % result)
            if uri is None:
                pasos.append(">> No hay destino para la foto.")
                aviso("Diagnóstico de la cámara", "\n".join(pasos))
                return
            img, copiados = self._leer_uri_imagen(uri)
            pasos.append("Bytes recibidos de la cámara: %s" % copiados)
            self._eliminar_uri(resolver, uri)
            if img is None:
                pasos.append(">> La cámara no dejó una foto utilizable.")
                pasos.append("Solución: usa el botón 'Cargar foto(s) y calificar'. "
                             "Toma la foto con tu cámara normal y luego selecciónala.")
                aviso("Diagnóstico de la cámara", "\n".join(pasos))
                return
            pasos.append("OpenCV leyó la imagen: sí (%dx%d)"
                         % (img.shape[1], img.shape[0]))
        except Exception as e:
            import traceback
            pasos.append(">> ERROR: %s" % e)
            pasos.append(traceback.format_exc())
            aviso("Diagnóstico de la cámara", "\n".join(pasos))
            return
        self._calificar_imagen(img)

    # --- Vía alterna: elegir una foto ya tomada desde la galería (muy confiable) ---
    _COD_GALERIA = 0xDA7A

    def _cargar_galeria_android(self):
        try:
            from jnius import autoclass  # type: ignore
            from android import activity  # type: ignore
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            ctx = PythonActivity.mActivity
            Intent = autoclass("android.content.Intent")
            intent = Intent(Intent.ACTION_GET_CONTENT)
            intent.setType("image/*")
            intent.putExtra(Intent.EXTRA_ALLOW_MULTIPLE, True)
            intent.addCategory(Intent.CATEGORY_OPENABLE)
            activity.bind(on_activity_result=self._on_galeria_result)
            ctx.startActivityForResult(intent, self._COD_GALERIA)
        except Exception as e:
            aviso("Cargar foto", "No se pudo abrir la galería: %s" % e)

    def _on_galeria_result(self, request, result, intent):
        if request != self._COD_GALERIA:
            return
        try:
            from android import activity  # type: ignore
            activity.unbind(on_activity_result=self._on_galeria_result)
        except Exception:
            pass
        if result not in (-1,) or intent is None:
            return
        uris = []
        try:
            clip = intent.getClipData()
            if clip is not None:
                for i in range(clip.getItemCount()):
                    uris.append(clip.getItemAt(i).getUri())
        except Exception:
            pass
        if not uris:
            try:
                d = intent.getData()
                if d is not None:
                    uris.append(d)
            except Exception:
                pass
        if not uris:
            return
        Clock.schedule_once(lambda dt: self._procesar_galeria_lote(uris), 0.1)

    def _procesar_galeria_lote(self, uris):
        items = []
        fallidos = 0
        for uri in uris:
            try:
                img, _ = self._leer_uri_imagen(uri)
            except Exception:
                img = None
            if img is None:
                fallidos += 1
                continue
            try:
                it = self._procesar_una(img)
            except Exception:
                it = None
            if it is None:
                fallidos += 1
            else:
                items.append(it)
        self.construir()
        self._mostrar_galeria(items, fallidos)

    def _foto_lista(self, ruta_foto):
        """Callback de plyer (iOS): lee la foto del archivo y la califica."""
        p = Path(str(ruta_foto)) if ruta_foto else None
        if not p or not p.exists():
            aviso("Cámara", "No se recibió la foto.")
            return
        import cv2
        img = cv2.imread(str(p))
        try:
            p.unlink()
        except Exception:
            pass
        if img is None:
            aviso("Cámara", "La foto no se pudo leer.")
            return
        self._calificar_imagen(img)

    def _procesar_una(self, img):
        """Califica UNA imagen, registra la nota y devuelve {textura, titulo}.
        Devuelve None si no se detectó la hoja (4 marcas), probando rotaciones."""
        import cv2
        import lector_omr as L
        clave_json = self._mapa_claves[self.sp_examen.text]
        datos = L.cargar_datos_examen(clave_json)
        # Si la foto viene girada, probamos las 4 orientaciones antes de rendirnos.
        marcas = None
        for rot in (None, cv2.ROTATE_90_CLOCKWISE, cv2.ROTATE_180,
                    cv2.ROTATE_90_COUNTERCLOCKWISE):
            prueba = img if rot is None else cv2.rotate(img, rot)
            gray = cv2.cvtColor(prueba, cv2.COLOR_BGR2GRAY)
            marcas = L.detectar_marcas(gray)
            if marcas is not None:
                img = prueba
                break
        if marcas is None:
            return None
        recta, scale = L.corregir_perspectiva(img, marcas)
        res = L.leer_respuestas(recta, datos.total_preguntas, scale)
        anotada = L.dibujar_anotaciones(recta, res, datos.clave, scale, datos.nombre)
        aciertos = sum(1 for q, r in res.respuestas.items()
                       if r == datos.clave.get(q))
        self.examen = datos.nombre
        nombre = self._nombre_de_codigo(res.codigo_estudiante)
        archivo = f"camara_{res.codigo_estudiante}"
        respuestas = {str(q): res.respuestas.get(q, "")
                      for q in range(1, datos.total_preguntas + 1)}
        existia = self._registrar_nota_camara(archivo, res.codigo_estudiante,
                                              aciertos, datos.total_preguntas,
                                              nombre, respuestas)
        self.resultados.append({"archivo": archivo, "codigo": res.codigo_estudiante,
                                "aciertos": aciertos, "total": datos.total_preguntas,
                                "nombre": nombre})
        if nombre:
            titulo = "%s  \u00b7  %d/%d" % (nombre, aciertos, datos.total_preguntas)
        else:
            titulo = "C\u00f3digo %s (sin nombre)  \u00b7  %d/%d" % (
                res.codigo_estudiante, aciertos, datos.total_preguntas)
        if existia:
            titulo += "  \u00b7  (nota actualizada)"
        return {"textura": self._np_a_textura(anotada), "titulo": titulo}

    def _calificar_imagen(self, img):
        """Califica una sola imagen (cámara) y la muestra en la galería."""
        try:
            item = self._procesar_una(img)
        except Exception as e:
            import traceback
            aviso("Error al procesar", "%s\n\n%s" % (e, traceback.format_exc()))
            return
        if item is None:
            aviso("No se detectó la hoja",
                  "No se vieron las 4 marcas de esquina. Repite la foto con "
                  "buena luz, la hoja completa, plana y sin sombras.")
            return
        self.construir()
        self._mostrar_galeria([item])

    def _mostrar_galeria(self, items, fallidos=0):
        """Muestra los resultados como galería: foto grande con flechas ‹ › a los
        costados para pasar de una hoja a la siguiente."""
        if not items:
            aviso("Calificar", "No se pudo leer ninguna hoja.%s\n\n"
                  "Repite las fotos con buena luz, la hoja completa y plana."
                  % (" (%d sin leer)" % fallidos if fallidos else ""))
            return
        estado = {"i": 0}
        cont = FloatLayout()
        img_w = Image(allow_stretch=True, keep_ratio=True,
                      size_hint=(0.80, 0.80), pos_hint={"center_x": 0.5, "center_y": 0.52})
        cont.add_widget(img_w)
        lbl = Label(markup=True, size_hint=(0.78, None), height=dp(30),
                    pos_hint={"center_x": 0.5, "top": 1},
                    halign="center", valign="middle")
        lbl.bind(size=lambda i, *_: setattr(i, "text_size", (i.width, i.height)))
        cont.add_widget(lbl)
        btn_prev = Factory.BotonAccion(text="\u2039", size_hint=(None, None),
                                       size=(dp(46), dp(66)),
                                       pos_hint={"x": 0.0, "center_y": 0.52})
        btn_next = Factory.BotonAccion(text="\u203a", size_hint=(None, None),
                                       size=(dp(46), dp(66)),
                                       pos_hint={"right": 1.0, "center_y": 0.52})
        cont.add_widget(btn_prev)
        cont.add_widget(btn_next)
        btn_x = Factory.BotonPeligro(text="X", size_hint=(None, None),
                                     size=(dp(40), dp(40)),
                                     pos_hint={"right": 1, "top": 1})
        cont.add_widget(btn_x)
        btn_ok = Factory.BotonExito(text="Confirmar", size_hint=(0.62, None),
                                    height=dp(50), pos_hint={"center_x": 0.5, "y": 0.015})
        cont.add_widget(btn_ok)

        def mostrar():
            i = estado["i"]
            it = items[i]
            img_w.texture = it["textura"]
            lbl.text = "%s   \u00b7   [b]%d/%d[/b]" % (it["titulo"], i + 1, len(items))
            uno = len(items) == 1
            btn_prev.opacity = 0 if uno else (0.3 if i == 0 else 1)
            btn_next.opacity = 0 if uno else (0.3 if i == len(items) - 1 else 1)
            btn_prev.disabled = uno or i == 0
            btn_next.disabled = uno or i == len(items) - 1

        def ir(d):
            estado["i"] = max(0, min(len(items) - 1, estado["i"] + d))
            mostrar()

        btn_prev.bind(on_release=lambda *_: ir(-1))
        btn_next.bind(on_release=lambda *_: ir(1))
        titulo_pop = "Resultados" + (" \u00b7 %d sin leer" % fallidos if fallidos else "")
        pop = Popup(title=titulo_pop, content=cont, size_hint=(0.97, 0.93))
        btn_x.bind(on_release=lambda *_: pop.dismiss())

        def _confirmar(*_):
            def _hacer():
                items.clear()      # libera las imágenes de memoria
                pop.dismiss()
            confirmar("Confirmar",
                      "Una vez confirmes ya no podr\u00e1s ver estas hojas de nuevo, "
                      "a menos que vuelvas a subir las im\u00e1genes.\n\n"
                      "Las notas quedan guardadas en el registro.",
                      _hacer, txt_si="S\u00ed, estoy seguro",
                      txt_no="D\u00e9jame revisar mejor")
        btn_ok.bind(on_release=_confirmar)
        mostrar()
        pop.open()

    def _nombre_de_codigo(self, codigo):
        """Busca el nombre del estudiante en la lista, a partir del código leído."""
        try:
            import registro_notas as RN
            app = App.get_running_app()
            cod = "".join(c for c in str(codigo) if c.isdigit()).zfill(3)
            for est in RN.cargar_roster(app.ruta_curso):
                if est.get("codigo") == cod:
                    return est.get("nombre", "")
        except Exception:
            pass
        return ""

    def _registrar_nota_camara(self, archivo, codigo, aciertos, total,
                               nombre="", respuestas=None):
        app = App.get_running_app()
        dir_salida = app.ruta_curso / "resultados"
        dir_salida.mkdir(parents=True, exist_ok=True)
        notas_path = dir_salida / f"notas_{self.examen}.json"
        registros = {}
        if notas_path.exists():
            try:
                for r in json.loads(notas_path.read_text(encoding="utf-8")):
                    registros[r.get("archivo", "")] = r
            except Exception:
                pass
        existia = archivo in registros
        registros[archivo] = {"archivo": archivo, "codigo": codigo,
                              "aciertos": aciertos, "total": total,
                              "nombre": nombre, "respuestas": respuestas or {}}
        notas_path.write_text(
            json.dumps(list(registros.values()), indent=2, ensure_ascii=False),
            encoding="utf-8")
        return existia

    @staticmethod
    def _np_a_textura(img_bgr):
        import cv2
        rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
        rgb = cv2.flip(rgb, 0)  # el origen de la textura de Kivy está abajo
        h, w = rgb.shape[:2]
        tex = Texture.create(size=(w, h), colorfmt="rgb")
        tex.blit_buffer(rgb.tobytes(), colorfmt="rgb", bufferfmt="ubyte")
        return tex

    def _popup_resultado(self, anotada, codigo, aciertos, total, nombre=""):
        cont = FloatLayout()
        img_w = Image(texture=self._np_a_textura(anotada),
                      allow_stretch=True, keep_ratio=True,
                      size_hint=(1, 1), pos_hint={"x": 0, "y": 0})
        cont.add_widget(img_w)
        btn_x = Factory.BotonPeligro(text="X", size_hint=(None, None),
                                     size=(dp(42), dp(42)),
                                     pos_hint={"right": 1, "top": 1})
        cont.add_widget(btn_x)
        if nombre:
            titulo = f"{nombre}  \u00b7  {aciertos}/{total}"
        else:
            titulo = f"C\u00f3digo {codigo} (sin nombre en lista)  \u00b7  {aciertos}/{total}"
        pop = Popup(title=titulo,
                    content=cont, size_hint=(0.96, 0.92))
        btn_x.bind(on_release=lambda *_: pop.dismiss())
        pop.open()


# ─────────────────────────────────────────────────────────────────────────────
# PANTALLA: REGISTRO DE NOTAS (Paso 5)  ·  Excel cruzando código -> estudiante
# ─────────────────────────────────────────────────────────────────────────────
class RegistroScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.cont = BoxLayout(orientation="vertical", padding=dp(14), spacing=dp(8))
        self.add_widget(self.cont)
        self.ultimo_excel = None
        self.mensaje = ""

    def on_pre_enter(self, *_):
        self.ultimo_excel = None
        self.mensaje = ""
        self.construir()

    def construir(self):
        app = App.get_running_app()
        self.cont.clear_widgets()
        b = self.cont
        b.add_widget(Encabezado(f"Registro de notas \u00b7 {app.grado}\u00b0{app.curso}"))

        import registro_notas as RN
        examenes = RN._examenes_disponibles(app.ruta_curso)
        if not examenes:
            b.add_widget(Label(text="Aún no hay exámenes calificados.\n"
                                    "Primero califica hojas en 'Calificar'."))
            self._volver(b)
            return

        fila = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
        fila.add_widget(Label(text="Examen:", size_hint_x=0.3))
        self.sp_examen = Spinner(text=examenes[0], values=examenes)
        fila.add_widget(self.sp_examen)
        b.add_widget(fila)

        # Período semestral para el envío al coordinador (formato AAAA-S)
        import datetime
        hoy = datetime.date.today()
        anios = [str(a) for a in range(hoy.year - 1, hoy.year + 2)]
        fila_per = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
        fila_per.add_widget(Label(text="Período:", size_hint_x=0.3))
        self.sp_anio = Spinner(text=str(hoy.year), values=anios, size_hint_x=0.4)
        fila_per.add_widget(self.sp_anio)
        self.sp_semestre = Spinner(text=("1" if hoy.month <= 6 else "2"),
                                   values=("1", "2"), size_hint_x=0.3)
        fila_per.add_widget(self.sp_semestre)
        b.add_widget(fila_per)

        btn = Factory.BotonAccion(text="Generar registro Excel", size_hint_y=None,
                                  height=dp(52))
        btn.bind(on_release=self._generar)
        b.add_widget(btn)

        self.lbl_msg = Label(text=self.mensaje, size_hint_y=None, height=dp(60))
        b.add_widget(self.lbl_msg)

        if self.ultimo_excel:
            btn_abrir = Button(text="Abrir Excel", size_hint_y=None, height=dp(44))
            btn_abrir.bind(on_release=lambda *_: self._abrir(self.ultimo_excel))
            b.add_widget(btn_abrir)

        btn_rep = Factory.BotonAccion(text="Reporte del docente (estadísticas)",
                                      size_hint_y=None, height=dp(48))
        btn_rep.bind(on_release=self._reporte_docente)
        b.add_widget(btn_rep)

        btn_coord = Factory.BotonOutline(text="Enviar planilla al coordinador",
                                         size_hint_y=None, height=dp(44))
        btn_coord.bind(on_release=self._enviar_coordinador)
        b.add_widget(btn_coord)

        b.add_widget(Label(text="", size_hint_y=1))
        self._volver(b)

    def _reporte_docente(self, *_):
        app = App.get_running_app()
        examen = self.sp_examen.text
        try:
            import reportes
            ruta = reportes.generar_reporte_docente(app.ruta_curso, examen)
        except Exception as e:
            aviso("Reporte del docente", f"No se pudo generar:\n{e}")
            return
        abrir_archivo_excel(ruta)

    def _enviar_coordinador(self, *_):
        app = App.get_running_app()
        examen = self.sp_examen.text
        periodo = f"{self.sp_anio.text}-{self.sp_semestre.text}"
        try:
            import reportes
            ruta = reportes.generar_json_salon(app.ruta_curso, examen, periodo)
        except Exception as e:
            aviso("Enviar al coordinador",
                  "json generado: false\n\n%s" % e)
            return
        aviso("Enviar al coordinador",
              "json generado: true\n\nArchivo:\n%s" % ruta.name)

    def _volver(self, b):
        btn = Button(text="← Volver al menú", size_hint_y=None, height=dp(44))
        btn.bind(on_release=lambda *_: setattr(App.get_running_app().sm, "current", "dashboard"))
        b.add_widget(btn)

    def _generar(self, *_):
        app = App.get_running_app()
        import registro_notas as RN
        examen = self.sp_examen.text
        try:
            ruta = RN.generar_registro(app.ruta_curso, examen)
        except Exception as e:
            aviso("Error", f"No se pudo generar el registro:\n{e}")
            return
        if ruta is None:
            aviso("Falta la planilla",
                  "No hay lista de estudiantes para este curso.\n"
                  "Créala en 'Estudiantes' y vuelve a intentar.")
            return
        self.ultimo_excel = str(ruta)
        self.mensaje = f"Registro generado:\n{ruta.name}"
        self.construir()

    def _abrir(self, ruta):
        abrir_archivo_excel(ruta)


# ─────────────────────────────────────────────────────────────────────────────
# APP
# ─────────────────────────────────────────────────────────────────────────────
class CalificadorApp(App):
    profesor = ""
    grado = ""
    curso = ""
    ruta_curso = None

    def tap_pulsar(self, widget):
        """Vibración corta al pulsar cualquier botón, salvo los de cerrar/salir."""
        if getattr(widget, "no_vibra", False):
            return
        _vibrar()

    def build(self):
        self.title = "CalificadorApp"
        # En Android la carpeta de instalación es de solo lectura: movemos el
        # "directorio de trabajo" a la carpeta privada de la app (sí escribible),
        # para que 'profesores/' y todos los datos se guarden sin problemas.
        if platform == "android":
            try:
                os.chdir(self.user_data_dir)
            except Exception:
                pass
        _registrar_fuentes()
        Window.clearcolor = FONDO
        if platform != "android":
            Window.size = (400, 720)   # vista tipo celular para probar en PC
        GS.inicializar_sistema()
        self.sm = ScreenManager(transition=SlideTransition(duration=0.18))
        self.sm.add_widget(SetupScreen(name="setup"))
        self.sm.add_widget(HomeScreen(name="home"))
        self.sm.add_widget(DashboardScreen(name="dashboard"))
        self.sm.add_widget(EstudiantesScreen(name="estudiantes"))
        self.sm.add_widget(EsamenesScreen(name="examenes"))
        self.sm.add_widget(ClavesScreen(name="claves"))
        self.sm.add_widget(CalificarScreen(name="calificar"))
        self.sm.add_widget(RegistroScreen(name="registro"))
        self.sm.current = "home" if _subdirs(BASE) else "setup"
        return self.sm

    def on_start(self):
        _solicitar_permisos_android()

    def ir_a_home(self):
        self.sm.current = "home"


if __name__ == "__main__":
    CalificadorApp().run()
