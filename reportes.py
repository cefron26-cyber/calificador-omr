# -*- coding: utf-8 -*-
"""
reportes.py — Salidas del Calificador OMR (lado profesor):

1) generar_json_salon(ruta_curso, examen)
   Crea UN archivo JSON por salón (todo 4°B en un archivo) con, por estudiante:
   código, número de lista, nombre, aciertos y las respuestas INCORRECTAS
   (qué pregunta y qué marcó). Incluye la clave del examen una sola vez.
   Es el archivo liviano que luego se envía al coordinador (vía Drive).

2) generar_reporte_docente(ruta_curso, examen)
   Crea un Excel SEPARADO (no la planilla de copiar/pegar) con estadísticas
   útiles para el docente: promedio, mediana, moda, nota más alta y quién,
   aprobados/reprobados, % de aprobación, análisis por pregunta y GRÁFICAS.
   Todo con FÓRMULAS de Excel (no valores quemados), para que recalcule solo.
"""
import json
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

import registro_notas as RN
import cripto

AZUL = "1F3864"
DORADO = "F2B705"
FUENTE = "Arial"


def _archivo_seguro(nombre):
    s = "".join(c if (c.isalnum() or c in " -_") else "_" for c in str(nombre))
    return s.strip().replace(" ", "_") or "examen"


def _buscar_clave(ruta_curso, examen):
    """Encuentra el JSON de clave cuyo campo 'examen' coincide."""
    carpeta = Path(ruta_curso) / "examenes_claves"
    if not carpeta.exists():
        return None
    for f in carpeta.glob("respuestas_*.json"):
        try:
            d = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        if d.get("examen") == examen:
            return d
    return None


# ─────────────────────────────────────────────────────────────────────────────
# 1) JSON POR SALÓN (para el coordinador)
# ─────────────────────────────────────────────────────────────────────────────
def generar_json_salon(ruta_curso, examen, periodo):
    ruta_curso = Path(ruta_curso)
    clave_data = _buscar_clave(ruta_curso, examen)
    if clave_data is None:
        raise RuntimeError("No se encontró la clave de ese examen.")
    arr = clave_data.get("clave_correctas", [])
    total = int(clave_data.get("total_preguntas", len(arr)))
    clave = {str(i + 1): arr[i] for i in range(len(arr))}

    roster = RN.cargar_roster(ruta_curso)                 # [{n_lista,codigo,nombre}]
    notas = RN.cargar_notas(ruta_curso, examen)           # registros con respuestas
    notas_por_cod = {str(n.get("codigo", "")).zfill(3): n for n in notas}

    grado = ruta_curso.parent.name
    curso = ruta_curso.name

    def incorrectas_de(reg):
        resp = reg.get("respuestas", {}) or {}
        fallos = []
        for q, correcta in clave.items():
            marco = resp.get(q, "")
            if marco != correcta:
                fallos.append({"p": int(q), "marco": marco})
        return fallos

    estudiantes = []
    usados = set()
    for est in roster:
        cod = str(est["codigo"]).zfill(3)
        usados.add(cod)
        reg = notas_por_cod.get(cod)
        if reg is None:
            estudiantes.append({
                "codigo": cod, "n_lista": est["n_lista"],
                "nombre": est["nombre"], "presento": False,
            })
        else:
            estudiantes.append({
                "codigo": cod, "n_lista": est["n_lista"],
                "nombre": est["nombre"], "presento": True,
                "aciertos": reg.get("aciertos"), "total": total,
                "incorrectas": incorrectas_de(reg),
            })

    # Hojas calificadas cuyo código no está en la lista (no se pierde nada)
    for cod, reg in notas_por_cod.items():
        if cod in usados:
            continue
        estudiantes.append({
            "codigo": cod, "n_lista": None,
            "nombre": reg.get("nombre", ""), "presento": True,
            "aciertos": reg.get("aciertos"), "total": total,
            "incorrectas": incorrectas_de(reg), "fuera_de_lista": True,
        })

    payload = {
        "version": 1,
        "grado": grado, "curso": curso, "examen": examen,
        "periodo": periodo,
        "total_preguntas": total, "clave": clave,
        "fecha": datetime.date.today().isoformat(),
        "estudiantes": estudiantes,
    }

    carpeta = ruta_curso / "resultados"
    carpeta.mkdir(parents=True, exist_ok=True)
    texto = json.dumps(payload, ensure_ascii=False, indent=2)
    destino = carpeta / f"coordinador_{grado}{curso}_{periodo}.json.enc"
    destino.write_bytes(cripto.cifrar(texto.encode("utf-8")))
    return destino


# ─────────────────────────────────────────────────────────────────────────────
# 2) REPORTE DEL DOCENTE (Excel con fórmulas + gráficas)
# ─────────────────────────────────────────────────────────────────────────────
def _fuente(c, size=11, bold=False, color="000000"):
    c.font = Font(name=FUENTE, size=size, bold=bold, color=color)


def generar_reporte_docente(ruta_curso, examen):
    ruta_curso = Path(ruta_curso)
    clave_data = _buscar_clave(ruta_curso, examen)
    if clave_data is None:
        raise RuntimeError("No se encontró la clave de ese examen.")
    arr = clave_data.get("clave_correctas", [])
    total = int(clave_data.get("total_preguntas", len(arr)))
    clave = {q + 1: arr[q] for q in range(len(arr))}

    roster = RN.cargar_roster(ruta_curso)
    notas = RN.cargar_notas(ruta_curso, examen)
    notas_por_cod = {str(n.get("codigo", "")).zfill(3): n for n in notas}

    grado = ruta_curso.parent.name
    curso = ruta_curso.name

    wb = Workbook()

    # ───── Hoja 1: Resumen ─────
    ws = wb.active
    ws.title = "Resumen"
    azul = PatternFill("solid", fgColor=AZUL)
    centro = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Reporte del docente — {examen}"
    _fuente(ws["A1"], size=14, bold=True)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{grado}°{curso}  ·  {total} preguntas  ·  {datetime.date.today().isoformat()}"
    _fuente(ws["A2"], size=10, color="555555")

    enc = ["N° Lista", "Código", "Nombre", "Aciertos", "Total", "Nota", "Estado"]
    fila_enc = 4
    for j, t in enumerate(enc, start=1):
        c = ws.cell(row=fila_enc, column=j, value=t)
        c.fill = azul
        _fuente(c, bold=True, color="FFFFFF")
        c.alignment = centro

    f0 = fila_enc + 1
    fila = f0
    for est in roster:
        cod = str(est["codigo"]).zfill(3)
        reg = notas_por_cod.get(cod)
        ws.cell(row=fila, column=1, value=est["n_lista"])
        ws.cell(row=fila, column=2, value=cod)
        ws.cell(row=fila, column=3, value=est["nombre"])
        if reg is not None:
            ws.cell(row=fila, column=4, value=reg.get("aciertos", 0))
            ws.cell(row=fila, column=5, value=total)
            # Nota = MAX(1, TRUNC(aciertos/total*10, 1))  → igual que la app
            ws.cell(row=fila, column=6,
                    value=f"=MAX(1,TRUNC(D{fila}/E{fila}*10,1))")
            ws.cell(row=fila, column=7,
                    value=f'=IF(F{fila}>=6,"Aprobó","Reprobó")')
        else:
            ws.cell(row=fila, column=5, value=total)
            ws.cell(row=fila, column=7, value="No presentó")
        for col in range(1, 8):
            _fuente(ws.cell(row=fila, column=col))
        fila += 1
    fN = fila - 1  # última fila de datos
    rango_nota = f"F{f0}:F{fN}" if fN >= f0 else f"F{f0}:F{f0}"
    rango_nom = f"C{f0}:C{fN}" if fN >= f0 else f"C{f0}:C{f0}"

    # Estadísticas (con fórmulas; IFERROR para no dejar errores)
    est_def = [
        ("Estadísticas", None),
        ("Presentaron", f"=COUNT({rango_nota})"),
        ("Promedio", f"=IFERROR(AVERAGE({rango_nota}),0)"),
        ("Mediana", f"=IFERROR(MEDIAN({rango_nota}),0)"),
        ("Moda", f'=IFERROR(MODE({rango_nota}),"—")'),
        ("Nota más alta", f"=IFERROR(MAX({rango_nota}),0)"),
        ("Quién (más alta)",
         f'=IFERROR(INDEX({rango_nom},MATCH(MAX({rango_nota}),{rango_nota},0)),"—")'),
        ("Nota más baja", f"=IFERROR(MIN({rango_nota}),0)"),
        ("Aprobados (≥6)", f'=COUNTIF({rango_nota},">=6")'),
        ("Reprobados (<6)", f'=COUNTIF({rango_nota},"<6")'),
        ("% Aprobación", None),  # se calcula con referencia abajo
    ]
    col_lbl, col_val = 9, 10  # I, J
    r = fila_enc
    fila_aprob = None
    for nombre_st, formula in est_def:
        cL = ws.cell(row=r, column=col_lbl, value=nombre_st)
        _fuente(cL, bold=(formula is None and nombre_st == "Estadísticas"))
        if nombre_st == "Aprobados (≥6)":
            fila_aprob = r
        if nombre_st == "Presentaron":
            fila_pres = r
        if formula is not None:
            cV = ws.cell(row=r, column=col_val, value=formula)
            _fuente(cV)
        r += 1
    fila_pct = r - 1
    cpct = ws.cell(row=fila_pct, column=col_val,
                   value=f"=IFERROR(J{fila_aprob}/J{fila_pres},0)")
    cpct.number_format = "0.0%"
    _fuente(cpct)

    # Tablita auxiliar para la gráfica de torta (aprobados/reprobados)
    rtor = fila_pct + 2
    ws.cell(row=rtor, column=col_lbl, value="Aprobados")
    ws.cell(row=rtor, column=col_val, value=f"=J{fila_aprob}")
    ws.cell(row=rtor + 1, column=col_lbl, value="Reprobados")
    ws.cell(row=rtor + 1, column=col_val, value=f"=J{fila_aprob+1}")
    for rr in (rtor, rtor + 1):
        _fuente(ws.cell(row=rr, column=col_lbl))
        _fuente(ws.cell(row=rr, column=col_val))

    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 12

    # Gráfica de barras: nota por estudiante
    if fN >= f0:
        bar = BarChart()
        bar.title = "Nota por estudiante"
        bar.type = "col"
        bar.y_axis.title = "Nota"
        datos = Reference(ws, min_col=6, min_row=fila_enc, max_row=fN)
        cats = Reference(ws, min_col=3, min_row=f0, max_row=fN)
        bar.add_data(datos, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 8
        bar.width = 18
        ws.add_chart(bar, f"A{fN + 3}")

        # Gráfica de torta: aprobados vs reprobados
        pie = PieChart()
        pie.title = "Aprobados vs Reprobados"
        dref = Reference(ws, min_col=col_val, min_row=rtor, max_row=rtor + 1)
        lref = Reference(ws, min_col=col_lbl, min_row=rtor, max_row=rtor + 1)
        pie.add_data(dref, titles_from_data=False)
        pie.set_categories(lref)
        pie.height = 8
        pie.width = 10
        ws.add_chart(pie, f"I{rtor + 4}")

    # ───── Hoja 2: Respuestas (matriz cruda; alimenta el análisis) ─────
    wr = wb.create_sheet("Respuestas")
    wr.cell(row=1, column=1, value="Código")
    wr.cell(row=1, column=2, value="Nombre")
    for q in range(1, total + 1):
        wr.cell(row=1, column=2 + q, value=f"P{q}")
    for j in range(1, total + 3):
        _fuente(wr.cell(row=1, column=j), bold=True)
    rr = 2
    presentes = 0
    for est in roster:
        cod = str(est["codigo"]).zfill(3)
        reg = notas_por_cod.get(cod)
        if reg is None:
            continue
        resp = reg.get("respuestas", {}) or {}
        wr.cell(row=rr, column=1, value=cod)
        wr.cell(row=rr, column=2, value=est["nombre"])
        for q in range(1, total + 1):
            wr.cell(row=rr, column=2 + q, value=resp.get(str(q), ""))
        rr += 1
        presentes += 1
    ult_resp = rr - 1

    # ───── Hoja 3: Por pregunta (análisis de ítems con fórmulas) ─────
    wp = wb.create_sheet("Por pregunta")
    cab = ["Pregunta", "Correcta", "Aciertos", "% Acierto",
           "A", "B", "C", "D", "Opción errónea más marcada"]
    for j, t in enumerate(cab, start=1):
        c = wp.cell(row=1, column=j, value=t)
        c.fill = azul
        _fuente(c, bold=True, color="FFFFFF")
        c.alignment = centro
    # columnas auxiliares (ocultas) para hallar el distractor más marcado
    aux0 = 11  # K
    for q in range(1, total + 1):
        rfp = q + 1
        col_letter = get_column_letter(2 + q)  # columna de Pq en 'Respuestas'
        rango = f"Respuestas!{col_letter}2:{col_letter}{ult_resp}" if ult_resp >= 2 \
            else f"Respuestas!{col_letter}2:{col_letter}2"
        wp.cell(row=rfp, column=1, value=q)
        wp.cell(row=rfp, column=2, value=clave.get(q, ""))   # correcta (dato)
        wp.cell(row=rfp, column=3, value=f"=COUNTIF({rango},B{rfp})")
        npres = max(1, presentes)
        wp.cell(row=rfp, column=4, value=f"=IFERROR(C{rfp}/{npres},0)")
        wp.cell(row=rfp, column=4).number_format = "0.0%"
        for k, letra in enumerate("ABCD"):
            wp.cell(row=rfp, column=5 + k,
                    value=f'=COUNTIF({rango},"{letra}")')
        # aux: conteo de cada opción salvo la correcta (la correcta = -1)
        for k, letra in enumerate("ABCD"):
            colaux = aux0 + k
            cl = get_column_letter(5 + k)  # E..H (conteos visibles)
            wp.cell(row=rfp, column=colaux,
                    value=f'=IF($B{rfp}="{letra}",-1,{cl}{rfp})')
        kref0 = get_column_letter(aux0)
        kref1 = get_column_letter(aux0 + 3)
        wp.cell(row=rfp, column=9,
                value=(f'=IFERROR(IF(MAX({kref0}{rfp}:{kref1}{rfp})<=0,"—",'
                       f'INDEX({{"A","B","C","D"}},'
                       f'MATCH(MAX({kref0}{rfp}:{kref1}{rfp}),'
                       f'{kref0}{rfp}:{kref1}{rfp},0))),"—")'))
        for j in range(1, 10):
            _fuente(wp.cell(row=rfp, column=j))
    # ocultar columnas auxiliares
    for k in range(4):
        wp.column_dimensions[get_column_letter(aux0 + k)].hidden = True
    wp.column_dimensions["I"].width = 26

    # Gráfica: % de acierto por pregunta (dificultad)
    if total >= 1:
        ch = BarChart()
        ch.title = "% de acierto por pregunta"
        ch.type = "col"
        ch.y_axis.title = "% acierto"
        d = Reference(wp, min_col=4, min_row=1, max_row=total + 1)
        cpre = Reference(wp, min_col=1, min_row=2, max_row=total + 1)
        ch.add_data(d, titles_from_data=True)
        ch.set_categories(cpre)
        ch.height = 8
        ch.width = 20
        wp.add_chart(ch, f"A{total + 3}")

    carpeta = ruta_curso / "resultados_excel"
    carpeta.mkdir(parents=True, exist_ok=True)
    destino = carpeta / f"reporte_{curso}_{_archivo_seguro(examen)}.xlsx"
    wb.save(destino)
    return destino
